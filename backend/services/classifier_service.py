import os
import pandas as pd
import requests
import json
import re
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
import time
import random
import difflib

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 기본 API 키 설정
DEFAULT_KAKAO_API_KEY = 'c3207609f232ab9972310f876e22e233'

# 주소 컬럼 후보
ADDRESS_CANDIDATES = [
    '받는분주소', '받는분 주소', '수취인주소', '수취인 주소', '주소', '배송지주소', '배송지 주소', 
    '수령인 주소(전체)', '수령인 주소', '수령인주소', '수령자 도로명 주소*', '수령자 도로명 주소'
]

# 컬럼 매핑
COLUMN_MAP = {
    '수령자': ['수령자', '수령인', '받는분', '수취인', '수취인명', '수령인명'],
    '수령자도로명주소': ['수령자 도로명 주소', '수령인 주소*', '수령인 주소(전체)', '받는분주소', '주소', '수령인 도로명 주소'],
    '수령인주소': ['수령자 도로명 주소', '수령인 주소*', '수령인 주소(전체)', '받는분주소', '주소', '수령인 도로명 주소'],
    '수령자상세주소': ['수령자 상세주소', '수령인 상세주소*', '상세주소', '상세 주소', '수령인상세주소', '받는분상세주소', '수령인 상세 주소', '수령자 상세 주소'],
    '수령인상세주소': ['수령자 상세주소', '수령인 상세주소*', '상세주소', '상세 주소', '수령인상세주소', '받는분상세주소', '수령인 상세 주소', '수령자 상세 주소'],
    '수령자연락처': ['수령자 연락처', '수령인 핸드폰*', '수령인 연락처', '수령인 휴대전화', '수취인연락처1', '수령자 연락처*'],
    '수령인연락처1': ['수령자 연락처', '수령인 핸드폰*', '수령인 연락처', '수령인 휴대전화', '수취인연락처1', '수령자 연락처*'],
    '상품명': ['상품명*', '상품명(총각네 쇼핑몰)', '상품명'],
    '거래처주문코드': ['거래처 주문번호', '거래처주문코드', '주문번호', '상품주문번호'],
    '주문자': ['주문자*', '주문자명', '주문자', '구매자명', '구매자', '보내는분'],
    '수량': ['수량'],
    '배송메시지': ['배송메시지', '비고2(배송메시지)', '출입방법', '출입 방법'],
}

def find_column(row, key):
    """row에서 key에 해당하는 컬럼의 값을 찾아 반환 (COLUMN_MAP 기반)"""
    candidates = COLUMN_MAP.get(key, [])
    row_keys = [str(k).strip() for k in row.keys()]
    
    def normalize(s):
        return re.sub(r'[^가-힣a-zA-Z0-9]', '', str(s)).lower()
    
    # 1. 완전일치
    for col in candidates:
        for rk in row_keys:
            if col == rk:
                return row.get(rk, '')
    
    # 2. normalize 일치
    for col in candidates:
        col_base = normalize(col)
        for rk in row_keys:
            rk_base = normalize(rk)
            if col_base == rk_base:
                return row.get(rk, '')
    
    # 3. 부분일치(가장 짧은 컬럼 우선)
    matches = []
    for col in candidates:
        col_base = normalize(col)
        for rk in row_keys:
            rk_base = normalize(rk)
            if col_base in rk_base or rk_base in col_base:
                matches.append((len(rk_base), rk, row.get(rk, '')))
    if matches:
        matches.sort()
        return matches[0][2]
    
    # 4. 직접 키 매칭 (fallback)
    if key in row:
        return row[key]
        
    # 5. 정규화된 키로 직접 매칭 (fallback)
    target = normalize(key)
    for col, val in row.items():
        if normalize(col) == target:
            return val
    
    # 6. 부분 매칭 (fallback)
    for col, val in row.items():
        if target in normalize(col):
            return val
    
    return ''

def clean_address_for_output(address):
    """출력용 주소 클린업"""
    if not isinstance(address, str):
        return str(address) if address else ''
    
    # '@' → '아파트'
    address = address.replace('@', '아파트')
    # 별모양 특수문자 → '*'
    for star in ['★', '☆', '✪', '✯', '✩', '✬', '✭', '✮', '✰', '✶', '✷', '✸', '✹', '✺', '✻', '✼', '✽', '✾', '✿', '❀', '❁', '❂', '❃', '❄', '❅', '❆', '❇', '❈', '❉', '❊', '❋']:
        address = address.replace(star, '*')
    # 로마숫자 → 숫자
    roman_map = {'Ⅰ': '1', 'Ⅱ': '2', 'Ⅲ': '3', 'Ⅳ': '4', 'Ⅴ': '5', 'Ⅵ': '6', 'Ⅶ': '7', 'Ⅷ': '8', 'Ⅸ': '9', 'Ⅹ': '10'}
    for k, v in roman_map.items():
        address = address.replace(k, v)
    # '--', 'ㅡ', '~' → '-', '&' → '앤'
    address = address.replace('--', '-').replace('ㅡ', '-').replace('~', '-').replace('&', '앤')
    return address

def replace_company_name(address):
    """회사명 치환"""
    if not isinstance(address, str):
        return address
    return address.replace('(주)자연의 모든것', '총각네')

def format_phone_number(phone):
    """전화번호 포맷팅"""
    if pd.isna(phone):
        return ''
    if not isinstance(phone, str):
        phone = str(phone)
    phone = phone.strip().replace('-', '')
    
    # 050x-xxxx-xxxx (12자리)
    if re.match(r'^050[0-9]', phone) and len(phone) == 12:
        return f"{phone[:4]}-{phone[4:8]}-{phone[8:]}"
    # 02-xxx-xxxx or 02-xxxx-xxxx
    if phone.startswith('02') and len(phone) in (9, 10):
        if len(phone) == 9:
            return f"{phone[:2]}-{phone[2:5]}-{phone[5:]}"
        else:
            return f"{phone[:2]}-{phone[2:6]}-{phone[6:]}"
    # 3자리 지역번호(031, 051 등)
    elif re.match(r'^0[3-6][0-9]', phone) and len(phone) in (10, 11):
        if len(phone) == 10:
            return f"{phone[:3]}-{phone[3:6]}-{phone[6:]}"
        else:
            return f"{phone[:3]}-{phone[3:7]}-{phone[7:]}"
    # 휴대폰(010, 011, 016, 017, 018, 019)
    elif re.match(r'^01[016789]', phone) and len(phone) == 11:
        return f"{phone[:3]}-{phone[3:7]}-{phone[7:]}"
    elif re.match(r'^01[016789]', phone) and len(phone) == 10:
        return f"{phone[:3]}-{phone[3:6]}-{phone[6:]}"
    # 1588, 1577, 16xx, 18xx 등 4-4
    elif re.match(r'^(15|16|18)[0-9]{2}', phone) and len(phone) == 8:
        return f"{phone[:4]}-{phone[4:]}"
    elif re.match(r'^070', phone) and len(phone) == 11:
        return f"{phone[:3]}-{phone[3:7]}-{phone[7:]}"
    
    return phone

def load_api_key():
    """API 키를 로드합니다."""
    api_key_file = os.path.join('data', 'api_key.txt')
    if os.path.exists(api_key_file):
        with open(api_key_file, 'r', encoding='utf-8') as f:
            return f.read().strip()
    return DEFAULT_KAKAO_API_KEY

def load_zip_codes(zip_type):
    """우편번호를 로드합니다."""
    if zip_type == 'day':
        file_path = os.path.join('data', 'system', '당일_우편번호.csv')
    elif zip_type == 'dawn':
        file_path = os.path.join('data', 'system', '새벽_우편번호.csv')
    else:
        return set()
    
    if os.path.exists(file_path):
        df = pd.read_csv(file_path, header=None)
        return {str(int(str(z).strip())).zfill(5) for z in df[0]}
    return set()

def get_zipcode_from_address(address, api_key):
    """주소에서 우편번호를 추출합니다."""
    if not address or not isinstance(address, str):
        return ""
    
    # 다양한 주소 형태로 시도
    tried = set()
    addr_try_list = []
    seen = set()
    
    # 1. 도로명+건물번호
    m = re.search(r'([가-힣A-Za-z0-9\s]+(로|길))\s*\d+[\-\d]*', address)
    if m:
        road_addr = m.group(0).strip()
        if road_addr and road_addr not in seen:
            addr_try_list.append(road_addr)
            seen.add(road_addr)
    
    # 2. 지번주소(동+번지)
    m2 = re.search(r'([가-힣A-Za-z0-9\s]+동)\s*\d+[\-\d]*', address)
    if m2:
        jibun_addr = m2.group(0).strip()
        if jibun_addr and jibun_addr not in seen:
            addr_try_list.append(jibun_addr)
            seen.add(jibun_addr)
    
    # 3. 원본 주소
    if address and address not in seen:
        addr_try_list.append(address)
        seen.add(address)
    
    # 4. 괄호제거
    bracketless = re.sub(r'\(.*?\)', '', address).strip()
    if bracketless and bracketless not in seen:
        addr_try_list.append(bracketless)
        seen.add(bracketless)
    
    # 5. 마지막 시도: 동/로까지만 남기고 API에 질의
    road_only = re.sub(r'\s+\d+.*', '', address)
    if road_only and road_only not in seen:
        addr_try_list.append(road_only)
        seen.add(road_only)
    
    for addr_try in addr_try_list:
        if not addr_try or addr_try in tried:
            continue
        tried.add(addr_try)
        
        zipcode = search_zipcode_kakao(addr_try, api_key)
        if zipcode:
            try:
                zc_str = str(int(zipcode)).zfill(5)
            except Exception:
                zc_str = zipcode
            return zc_str
    
    return ""

def search_zipcode_kakao(address, api_key):
    """카카오 API로 우편번호 검색"""
    max_retries = 3
    retry_delay = 1
    
    for attempt in range(max_retries):
        try:
            time.sleep(0.1 + random.uniform(0, 0.2))
            
            url = "https://dapi.kakao.com/v2/local/search/address.json"
            headers = {"Authorization": f"KakaoAK {api_key}"}
            params = {"query": address}
            
            response = requests.get(url, headers=headers, params=params)
            
            if response.status_code == 200:
                data = response.json()
                if data['documents']:
                    doc = data['documents'][0]
                    if doc.get('road_address') and doc['road_address'].get('zone_no'):
                        return doc['road_address']['zone_no']
                    elif doc.get('address') and doc['address'].get('zone_no'):
                        return doc['address']['zone_no']
                else:
                    return ""
            elif response.status_code == 429:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (attempt + 1))
                    continue
            else:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    continue
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
                continue
    return ""

def classify_by_zipcode(zipcode, day_zip_codes, dawn_zip_codes, row_dict, idx, classify_mode="all"):
    """우편번호를 기반으로 배송 유형을 분류"""
    try:
        if not zipcode:
            logger.warning(f"[{idx}] 우편번호가 없음")
            row_dict['delivery_type'] = '미분류'
            return {'type': 'unclassified', 'data': row_dict}
        
        zipcode_str = str(zipcode).zfill(5)
        
        # 1. 당일배송 확인 (우선순위 1위)
        if zipcode_str in day_zip_codes:
            logger.info(f"[{idx}] 당일배송 분류 (우편번호: {zipcode_str})")
            row_dict['delivery_type'] = '당일'
            return {'type': 'day', 'data': row_dict}
        
        # 2. 새벽배송 확인 (우선순위 2위)
        if zipcode_str in dawn_zip_codes:
            logger.info(f"[{idx}] 새벽배송 분류 (우편번호: {zipcode_str})")
            row_dict['delivery_type'] = '새벽'
            return {'type': 'dawn', 'data': row_dict}
        
        # 3. 택배배송 (우선순위 3위, 정상 배송)
        logger.info(f"[{idx}] 택배배송 분류 (우편번호: {zipcode_str})")
        row_dict['delivery_type'] = '택배'
        return {'type': 'delivery', 'data': row_dict}
        
    except Exception as e:
        logger.error(f"[{idx}] 분류 중 오류: {e}")
        row_dict['delivery_type'] = '미분류'
        return {'type': 'unclassified', 'data': row_dict}

def normalize_addr(addr):
    """주소 정규화 (중복 체크용)"""
    s = str(addr or '').strip()
    if s == '' or s.lower() == 'nan' or s.lower() == 'none' or s == '0' or s == '0 / 0':
        return ''
    # 모든 공백, 줄바꿈, 탭, 유니코드 공백 제거
    s = re.sub(r'[\s\r\n\t\u200b\u3000]', '', s)
    # 괄호 및 괄호 안 내용 제거
    s = re.sub(r'\(.*?\)', '', s)
    # 특수문자 제거(한글, 영문, 숫자만 남김)
    s = re.sub(r'[^가-힣a-zA-Z0-9]', '', s)
    return s.lower()

def save_with_yellow_header(df, output_path):
    """노란색 헤더와 파스텔톤 색상을 적용하여 저장"""
    df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 노란색 헤더
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws[1]:
        cell.fill = yellow_fill
    
    # 주소 컬럼 찾기
    address_candidates = [
        '받는분주소', '받는분 주소', '수취인주소', '수취인 주소', '수령인 주소*', '수령인주소', '수령인 주소',
        '배송지주소', '배송지 주소', '수령자 도로명 주소*', '송하인주소', '주소', '수령자 주소',
        '수령인 도로명 주소', '수령자도로명주소', '수령인도로명주소', '받는분 도로명 주소', '받는분도로명주소',
        '배송지 도로명 주소', '배송지도로명주소'
    ]
    
    address_col = None
    # 0. '받는분주소'가 있으면 무조건 address_col로 지정
    for i, cell in enumerate(ws[1], 1):
        val = str(cell.value or '').strip()
        if val == '받는분주소':
            address_col = i
            break
    
    if address_col is None:
        # 다른 주소 컬럼 찾기
        for i, cell in enumerate(ws[1], 1):
            val = str(cell.value or '').strip()
            if '주소' in val:
                address_col = i
                break
    
    # 파스텔톤 색상 적용
    if address_col is not None:
        addr_to_rows = {}
        for row in range(2, ws.max_row + 1):
            addr = str(ws.cell(row=row, column=address_col).value or '').strip()
            norm_addr = normalize_addr(addr)
            if norm_addr == '':
                continue
            if norm_addr not in addr_to_rows:
                addr_to_rows[norm_addr] = []
            addr_to_rows[norm_addr].append(row)
        
        if len(addr_to_rows) >= 2:
            colors = [
                "FFFF99", "FFEBEE", "E3F2FD", "E8F5E9", "FFFDE7", "F3E5F5", "FBE9E7",
                "E0F2F1", "FFF3E0", "F9FBE7", "EDE7F6", "F1F8E9", "FCE4EC"
            ]
            color_map = {}
            color_idx = 0
            for addr, rows in addr_to_rows.items():
                if len(rows) < 2:
                    continue
                color_map[addr] = PatternFill(start_color=colors[color_idx % len(colors)], end_color=colors[color_idx % len(colors)], fill_type='solid')
                color_idx += 1
            
            for addr, rows in addr_to_rows.items():
                if addr not in color_map:
                    continue
                fill = color_map[addr]
                for row in rows:
                    for cell in ws[row]:
                        cell.fill = fill
        
        # 헤더 노란색 다시 적용
        for cell in ws[1]:
            cell.fill = yellow_fill
    
    wb.save(output_path)

def get_incremented_filename(base, ext, output_dir):
    """중복되지 않는 파일명 생성"""
    if base.endswith(f'.{ext}'):
        base = base[:-(len(ext)+1)]
    filename = f"{base}.{ext}"
    full_path = os.path.join(output_dir, filename)
    i = 1
    while os.path.exists(full_path):
        filename = f"{base}({i}).{ext}"
        full_path = os.path.join(output_dir, filename)
        i += 1
    return full_path

def has_jonggakne_company(addr):
    """총각네 업체명 검증 함수 (GUI와 동일)"""
    if not isinstance(addr, str):
        return False
    import re
    # (주), ㈜, 공백 등 다양한 변형 매칭
    return bool(re.search(r'(\(주\)|㈜)\s*자연의\s*모든것', addr))

def has_nakgeun_text(addr):
    """이낙근 업체명 검증 함수 (GUI와 동일)"""
    if not isinstance(addr, str):
        return False
    return '이낙근' in addr

def has_gooknchan_text(addr):
    """국앤찬 업체명 검증 함수 (GUI와 동일)"""
    if not isinstance(addr, str):
        return False
    return '국앤찬' in addr

def make_dawn_product_name(row, chasu=None):
    """새벽배송 상품명 생성 (GUI와 동일한 업체명 로직 포함)"""
    # chasu가 '차수없음'이거나 None/빈값이면 상품명에 차수 미포함
    if chasu in [None, '', '차수없음']:
        chasu_str = ''
    else:
        chasu_str = f" {chasu}"
    
    buyer = find_column(row, '구매자명') or find_column(row, '주문자')
    product = find_column(row, '상품명')
    
    # 주문자명 마스킹 처리
    masked_buyer = ''
    if buyer and str(buyer).lower() != 'nan' and buyer is not None:
        buyer_str = str(buyer)
        if len(buyer_str) == 2:
            masked_buyer = buyer_str[0] + '*'
        elif len(buyer_str) >= 3:
            masked_buyer = buyer_str[0] + '*' * (len(buyer_str) - 2) + buyer_str[-1]
        else:
            masked_buyer = buyer_str
    else:
        masked_buyer = ''
    
    prefix = "주문 : " + masked_buyer
    
    # 업체명 확인을 위한 컬럼들 (주소, 주문자, 거래처주문코드 등)
    order_addr = find_column(row, '주문자주소') or find_column(row, '구매자주소') or find_column(row, '송하인주소')
    order_code = find_column(row, '거래처주문코드') or find_column(row, '상품주문번호')
    
    def safe_str(val):
        return str(val) if val is not None else ''
    
    # 주소에서 업체명 확인 (GUI와 동일한 로직)
    if order_addr:
        cleaned_addr = re.sub(r'[\s\-_]+', '', safe_str(order_addr))
        if has_nakgeun_text(order_addr):
            prefix += " 이낙근"
        if has_gooknchan_text(order_addr):
            prefix += " 국앤찬"
        if has_jonggakne_company(order_addr):  # GUI와 동일한 함수 사용
            prefix += " 총각네"
        if '맛사랑' in cleaned_addr:
            prefix += " 맛사랑"
        if '새농' in cleaned_addr:
            prefix += " 새농"
        if '당일장터' in cleaned_addr:
            prefix += " 당일장터"
        if '남촌상회' in cleaned_addr or '남촌과일' in cleaned_addr:
            prefix += " 남촌상회"
    
    # 거래처주문코드에서 업체명 확인
    if order_code and ('남촌상회' in safe_str(order_code) or '남촌과일' in safe_str(order_code)):
        prefix += " 남촌상회"
    
    # 구매자명/주문자명에서 업체명 확인
    if buyer and ('남촌상회' in safe_str(buyer) or '남촌과일' in safe_str(buyer)):
        prefix += " 남촌상회"
    
    # (주)봉봉후르츠 제거 로직 삭제 - GUI에는 없는 불필요한 로직이었음
    result = f"{prefix}{chasu_str} {product}"
    
    return result.strip()

def classify_delivery(file_path, address_col=None, detail_col=None, classify_mode="all", 
                     custom_dawn_zips=None, custom_day_zips=None, chasu=None, 
                     dawn_type='배송대행', sms_type='즉시전송', progress_callback=None):
    """배송 분류 메인 함수"""
    
    # 데이터 로드
    df = pd.read_excel(file_path, dtype=str)
    original_count = len(df)
    logger.info(f"데이터 로드 완료: {original_count}행")
    
    # API 키 로드
    api_key = load_api_key()
    
    # 분류 모드에 따라 필요한 우편번호만 로드
    if classify_mode == "day":
        # 당일배송만 선택 시 당일 우편번호만 로드
        day_zip_codes = custom_day_zips or load_zip_codes('day')
        dawn_zip_codes = set()  # 빈 set으로 설정
        logger.info(f"당일배송만 분류 - 당일배송 우편번호: {len(day_zip_codes)}개")
    elif classify_mode == "dawn":
        # 새벽배송만 선택 시 새벽 우편번호만 로드
        day_zip_codes = set()  # 빈 set으로 설정
        dawn_zip_codes = custom_dawn_zips or load_zip_codes('dawn')
        logger.info(f"새벽배송만 분류 - 새벽배송 우편번호: {len(dawn_zip_codes)}개")
    else:
        # 전체 분류 시 모든 우편번호 로드
        day_zip_codes = custom_day_zips or load_zip_codes('day')
        dawn_zip_codes = custom_dawn_zips or load_zip_codes('dawn')
        logger.info(f"당일배송 우편번호: {len(day_zip_codes)}개")
        logger.info(f"새벽배송 우편번호: {len(dawn_zip_codes)}개")
    
    # 주소 컬럼 자동 선택
    if not address_col:
        for col in ADDRESS_CANDIDATES:
            if col in df.columns:
                address_col = col
                break
    
    if not address_col:
        raise ValueError("주소 컬럼을 찾을 수 없습니다.")
    
    # 주소와 상세주소 합치기
    if detail_col and detail_col in df.columns:
        def make_full_addr(row):
            base = str(row[address_col]) if pd.notna(row[address_col]) else ''
            detail = str(row[detail_col]) if pd.notna(row[detail_col]) and str(row[detail_col]).strip() else ''
            base = base.rstrip(', ').strip()
            # '0 / 0' 처리 추가
            if detail and detail.lower() != 'nan' and detail != '0 / 0':
                return f"{base} {detail}".strip()
            else:
                return base
        df['최종주소'] = df.apply(make_full_addr, axis=1)
        address_for_classification = '최종주소'
    else:
        address_for_classification = address_col
    
    # 분류 결과 저장용
    day_results = []
    dawn_results = []
    delivery_results = []
    unclassified_results = []
    
    total_rows = len(df)
    processed = 0
    
    for idx, row in df.iterrows():
        if progress_callback:
            progress = (processed / total_rows) * 100
            progress_callback(progress)
        
        address = row[address_for_classification]
        
        # 위례 지역 새벽배송 자동 분류
        if '위례' in str(address):
            row_dict = row.to_dict()
            row_dict['delivery_type'] = '새벽'
            dawn_results.append(row_dict)
            processed += 1
            continue
        
        # 우편번호 추출
        zipcode = get_zipcode_from_address(address, api_key)
        
        # 행 데이터를 딕셔너리로 변환
        row_dict = row.to_dict()
        
        # 분류 실행
        if zipcode:
            row_classified = classify_by_zipcode(zipcode, day_zip_codes, dawn_zip_codes, row_dict, idx, classify_mode)
            
            # 분류 결과 저장
            if row_classified['type'] == 'day':
                day_results.append(row_classified['data'])
            elif row_classified['type'] == 'dawn':
                dawn_results.append(row_classified['data'])
            elif row_classified['type'] == 'delivery':
                delivery_results.append(row_classified['data'])
            else:
                unclassified_results.append(row_classified['data'])
        else:
            # 우편번호를 찾을 수 없는 경우 미분류
            row_dict['delivery_type'] = '미분류'
            unclassified_results.append(row_dict)
        
        processed += 1
    
    if progress_callback:
        progress_callback(100)
    
    # 결과 파일 저장 (사용자 데이터 디렉토리에 저장)
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "user")
    os.makedirs(output_dir, exist_ok=True)
    saved_files = []
    
    # 실제 분류된 행 수 계산 (수량 확장 후)
    actual_day_count = 0
    actual_dawn_count = 0
    actual_delivery_count = 0
    actual_unclassified_count = 0
    
    # 분류 모드에 따른 처리
    if classify_mode == "all":
        # 전체 분류
        
        # 당일배송
        if day_results:
            day_df = process_day_delivery(day_results, address_col, detail_col)
            actual_day_count = len(day_df)
            day_path = get_incremented_filename("당일배송_주문리스트", "xlsx", output_dir)
            save_with_yellow_header(day_df, day_path)
            saved_files.append(day_path)
        
        # 새벽배송
        if dawn_results:
            dawn_df = process_dawn_delivery(dawn_results, chasu, dawn_type, sms_type)
            actual_dawn_count = len(dawn_df)
            dawn_path = get_incremented_filename("새벽배송_주문리스트", "xlsx", output_dir)
            save_with_yellow_header(dawn_df, dawn_path)
            saved_files.append(dawn_path)
        
        # 택배배송
        if delivery_results:
            delivery_df = process_day_delivery(delivery_results, address_col, detail_col)
            actual_delivery_count = len(delivery_df)
            delivery_path = get_incremented_filename("택배배송_주문리스트", "xlsx", output_dir)
            save_with_yellow_header(delivery_df, delivery_path)
            saved_files.append(delivery_path)
        
        # 미분류 (주소오류)
        if unclassified_results:
            unclassified_df = pd.DataFrame(unclassified_results)
            actual_unclassified_count = len(unclassified_df)
            # delivery_type 컬럼 제거
            if 'delivery_type' in unclassified_df.columns:
                unclassified_df = unclassified_df.drop(columns=['delivery_type'])
            unclassified_path = get_incremented_filename("주소오류_미분류", "xlsx", output_dir)
            save_with_yellow_header(unclassified_df, unclassified_path)
            saved_files.append(unclassified_path)
    
    elif classify_mode == "day":
        # 당일배송만
        if day_results:
            day_df = process_day_delivery(day_results, address_col, detail_col)
            actual_day_count = len(day_df)
            day_path = get_incremented_filename("당일배송_주문리스트(단일)", "xlsx", output_dir)
            save_with_yellow_header(day_df, day_path)
            saved_files.append(day_path)
        
        # 미분류 저장 (당일배송과 동일한 처리 적용)
        other_results = dawn_results + delivery_results + unclassified_results
        if other_results:
            other_df = process_unclassified_data(other_results, address_col, detail_col)
            actual_unclassified_count = len(other_df)
            other_path = get_incremented_filename("미분류_주문리스트(당일불가지역)", "xlsx", output_dir)
            save_with_yellow_header(other_df, other_path)
            saved_files.append(other_path)
    
    elif classify_mode == "dawn":
        # 새벽배송만
        if dawn_results:
            dawn_df = process_dawn_delivery(dawn_results, chasu, dawn_type, sms_type)
            actual_dawn_count = len(dawn_df)
            dawn_path = get_incremented_filename("새벽배송_주문리스트(단일)", "xlsx", output_dir)
            save_with_yellow_header(dawn_df, dawn_path)
            saved_files.append(dawn_path)
        
        # 미분류 저장 (당일배송과 동일한 처리 적용)
        other_results = day_results + delivery_results + unclassified_results
        if other_results:
            other_df = process_unclassified_data(other_results, address_col, detail_col)
            actual_unclassified_count = len(other_df)
            other_path = get_incremented_filename("미분류_주문리스트", "xlsx", output_dir)
            save_with_yellow_header(other_df, other_path)
            saved_files.append(other_path)
    
    # 분류 결과 반환 (수량 확장 후 실제 행 수)
    results = {
        'total': actual_day_count + actual_dawn_count + actual_delivery_count + actual_unclassified_count,
        'day_count': actual_day_count,
        'dawn_count': actual_dawn_count,
        'delivery_count': actual_delivery_count,
        'unclassified_count': actual_unclassified_count,
        'saved_files': saved_files
    }
    
    logger.info(f"분류 완료 - 원본:{original_count}행 → 총:{results['total']}행 (당일:{results['day_count']} 새벽:{results['dawn_count']} 택배:{results['delivery_count']} 미분류:{results['unclassified_count']})")
    
    return results

def process_day_delivery(day_results, address_col, detail_col=None):
    """당일배송 데이터 처리"""
    df = pd.DataFrame(day_results)
    
    print(f"DEBUG: 처리 전 컬럼들: {list(df.columns)}")
    print(f"DEBUG: address_col: {address_col}, detail_col: {detail_col}")
    
    # GUI와 동일한 최종주소 생성 로직
    if detail_col and detail_col in df.columns and address_col in df.columns:
        print("DEBUG: 최종주소 컬럼 생성 시작")
        def make_full_addr(row):
            base = str(row[address_col]) if pd.notna(row[address_col]) else ''
            detail = str(row[detail_col]) if pd.notna(row[detail_col]) and str(row[detail_col]).strip() else ''
            base = base.rstrip(', ').strip()
            # '0 / 0' 처리 추가
            if detail and detail.lower() != 'nan' and detail != '0 / 0':
                return f"{base} {detail}".strip()
            else:
                return base
        df['최종주소'] = df.apply(make_full_addr, axis=1)
        print("DEBUG: 최종주소 컬럼 생성 완료")
    
    # 수량 반복 처리
    expanded_rows = []
    for _, row in df.iterrows():
        qty = 1
        qty_val = find_column(row, '수량')
        if qty_val:
            try:
                qty = int(qty_val)
            except:
                qty = 1
        
        for _ in range(qty):
            new_row = row.copy()
            # 수량 컬럼 찾아서 1로 설정
            for col in df.columns:
                if '수량' in col:
                    new_row[col] = '1'
            expanded_rows.append(new_row)
    
    df = pd.DataFrame(expanded_rows)
    
    print(f"DEBUG: 수량 처리 후 컬럼들: {list(df.columns)}")
    print(f"DEBUG: '최종주소' 컬럼 존재 여부: {'최종주소' in df.columns}")
    
    # GUI와 100% 동일한 주소 처리 로직
    if '최종주소' in df.columns:
        print("DEBUG: 최종주소 컬럼 사용")
        # 최종주소를 받는분주소에 복사
        df[address_col] = df['최종주소'].apply(clean_address_for_output)
        # 상세주소 컬럼들을 빈값으로 설정 (COLUMN_MAP 기반)
        for col in COLUMN_MAP['수령인상세주소']:
            if col in df.columns:
                print(f"DEBUG: 상세주소 컬럼 '{col}' 빈값으로 설정")
                df[col] = ''
        # 최종주소 컬럼 제거
        df = df.drop(columns=['최종주소'])
        print("DEBUG: 최종주소 컬럼 제거 완료")
    elif address_col in df.columns:
        print("DEBUG: 원본 주소 컬럼만 정리")
        df[address_col] = df[address_col].apply(clean_address_for_output)
    
    # 상세주소 컬럼들 삭제 (당일배송은 상세주소 컬럼 자체를 제거)
    detail_cols_to_remove = []
    
    # 1. COLUMN_MAP 기반 삭제
    for col in COLUMN_MAP['수령인상세주소']:
        if col in df.columns:
            detail_cols_to_remove.append(col)
            print(f"DEBUG: COLUMN_MAP 기반 삭제 대상: '{col}'")
    
    # 2. 추가 패턴으로 상세주소 컬럼 찾기
    for col in df.columns:
        col_lower = col.lower()
        if any(keyword in col_lower for keyword in ['상세주소', '상세 주소', 'detail']):
            if col not in detail_cols_to_remove:
                detail_cols_to_remove.append(col)
                print(f"DEBUG: 패턴 기반 삭제 대상: '{col}'")
    
    # 상세주소 컬럼들 삭제
    if detail_cols_to_remove:
        print(f"DEBUG: 삭제할 상세주소 컬럼들: {detail_cols_to_remove}")
        try:
            columns_before = set(df.columns)
            df = df.drop(columns=detail_cols_to_remove, errors='ignore')
            columns_after = set(df.columns)
            deleted_cols = columns_before - columns_after
            print(f"DEBUG: 실제 삭제된 컬럼들: {list(deleted_cols)}")
            print(f"DEBUG: 상세주소 컬럼 삭제 완료: {len(deleted_cols) > 0}")
        except Exception as e:
            print(f"DEBUG: 상세주소 컬럼 삭제 중 오류: {e}")
            print(f"DEBUG: 상세주소 컬럼 삭제 완료: False")
    else:
        print("DEBUG: 삭제할 상세주소 컬럼 없음")
    
    # delivery_type 컬럼 제거
    if 'delivery_type' in df.columns:
        df = df.drop(columns=['delivery_type'])
    
    print(f"DEBUG: 처리 후 컬럼들: {list(df.columns)}")
    
    # 전화번호 포맷팅 (원래 로직 유지)
    for col in COLUMN_MAP['수령인연락처1']:
        if col in df.columns:
            df[col] = df[col].apply(format_phone_number)
    
    return df

def process_dawn_delivery(dawn_results, chasu=None, dawn_type='배송대행', sms_type='즉시전송'):
    """새벽배송 데이터 처리"""
    template_path = os.path.join('data', 'system', '새벽배송양식.xlsx')
    if not os.path.exists(template_path):
        # 템플릿이 없으면 기본 처리
        df = pd.DataFrame(dawn_results)
        # delivery_type 컬럼 제거
        if 'delivery_type' in df.columns:
            df = df.drop(columns=['delivery_type'])
        return df
    
    template = pd.read_excel(template_path)
    columns = list(template.columns)
    tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    
    result = []
    for row_dict in dawn_results:
        qty = 1
        qty_val = find_column(row_dict, '수량')
        if qty_val:
            try:
                qty = int(qty_val)
            except:
                qty = 1
        
        for _ in range(qty):
            new_row = {col: '' for col in columns}
            new_row['배송요청일*'] = tomorrow
            new_row['거래처 주문번호'] = find_column(row_dict, '거래처주문코드')
            new_row['주문자*'] = find_column(row_dict, '주문자')
            new_row['수령자*'] = find_column(row_dict, '수령자')
            new_row['상품명*'] = make_dawn_product_name(row_dict, chasu)
            
            # 주소 합치기 - GUI와 동일한 로직
            if '최종주소' in row_dict and row_dict['최종주소']:
                # 이미 합쳐진 주소 사용
                full_addr = clean_address_for_output(replace_company_name(row_dict['최종주소']))
            else:
                # 수동으로 합치기
                road_addr = clean_address_for_output(replace_company_name(find_column(row_dict, '수령자도로명주소')))
                detail_addr = find_column(row_dict, '수령자상세주소')
                if detail_addr and str(detail_addr).strip() and str(detail_addr).strip().lower() != 'nan':
                    full_addr = f"{road_addr} {str(detail_addr).strip()}"
                else:
                    full_addr = road_addr
            
            new_row['수령자 도로명 주소*'] = full_addr.strip()
            new_row['수령자  상세 주소'] = ''  # 공백 두 개, 빈 값으로 설정 (주소가 합쳐졌으므로)
            new_row['수령자 연락처*'] = format_phone_number(find_column(row_dict, '수령자연락처'))
            
            dawn_msg = find_column(row_dict, '배송메시지')
            
            def is_etc_condition(val):
                if val is None:
                    return True
                sval = str(val).strip()
                if sval == '' or sval.lower() == 'nan' or sval == '0 / 0':
                    return True
                return False
            
            for col in columns:
                if '배송 받을 장소' in col and '상세' not in col:
                    new_row[col] = '기타' if is_etc_condition(dawn_msg) else '문 앞'
                if '배송 받을 장소 상세' in col:
                    new_row[col] = '문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송'
                if '배송 유형' in col:
                    new_row[col] = dawn_type or '배송대행'
                if '배송 문자 전송 시점' in col:
                    new_row[col] = sms_type or '즉시전송'
                if '출입방법' in col.replace(' ', '') or '출입 방법' in col:
                    # '0 / 0' 등 기타 조건일 때는 빈값으로 처리
                    new_row[col] = '' if is_etc_condition(dawn_msg) else dawn_msg
            
            result.append(new_row)
    
    return pd.DataFrame(result, columns=columns)

def process_unclassified_data(unclassified_results, address_col, detail_col=None):
    """미분류/주소오류 데이터 처리 (당일배송과 동일한 로직 적용)"""
    if not unclassified_results:
        return pd.DataFrame()
    
    df = pd.DataFrame(unclassified_results)
    
    # 1. 주소 합치기 로직 (당일배송과 동일)
    if detail_col and detail_col in df.columns and address_col in df.columns:
        def make_full_addr(row):
            base = str(row[address_col]) if pd.notna(row[address_col]) else ''
            detail = str(row[detail_col]) if pd.notna(row[detail_col]) and str(row[detail_col]).strip() else ''
            base = base.rstrip(', ').strip()
            # '0 / 0' 처리 추가
            if detail and detail.lower() != 'nan' and detail != '0 / 0':
                return f"{base} {detail}".strip()
            else:
                return base
        df['최종주소'] = df.apply(make_full_addr, axis=1)
    
    # 2. 수량 반복 처리
    expanded_rows = []
    for _, row in df.iterrows():
        qty = 1
        qty_val = find_column(row, '수량')
        if qty_val:
            try:
                qty = int(qty_val)
            except:
                qty = 1
        
        for _ in range(qty):
            new_row = row.copy()
            # 수량 컬럼 찾아서 1로 설정
            for col in df.columns:
                if '수량' in col:
                    new_row[col] = '1'
            expanded_rows.append(new_row)
    
    df = pd.DataFrame(expanded_rows)
    
    # 3. 주소 처리 로직 (당일배송과 동일)
    if '최종주소' in df.columns:
        # 최종주소를 받는분주소에 복사
        df[address_col] = df['최종주소'].apply(clean_address_for_output)
        # 상세주소 컬럼들을 빈값으로 설정
        for col in COLUMN_MAP['수령인상세주소']:
            if col in df.columns:
                df[col] = ''
        # 최종주소 컬럼 제거
        df = df.drop(columns=['최종주소'])
    elif address_col in df.columns:
        df[address_col] = df[address_col].apply(clean_address_for_output)
    
    # 4. 상세주소 컬럼들 삭제
    detail_cols_to_remove = []
    for col in COLUMN_MAP['수령인상세주소']:
        if col in df.columns:
            detail_cols_to_remove.append(col)
    
    # 추가 패턴으로 상세주소 컬럼 찾기
    for col in df.columns:
        col_lower = col.lower()
        if any(keyword in col_lower for keyword in ['상세주소', '상세 주소', 'detail']):
            if col not in detail_cols_to_remove:
                detail_cols_to_remove.append(col)
    
    if detail_cols_to_remove:
        df = df.drop(columns=detail_cols_to_remove, errors='ignore')
    
    # 5. delivery_type 컬럼 제거
    if 'delivery_type' in df.columns:
        df = df.drop(columns=['delivery_type'])
    
    # 6. 전화번호 포맷팅
    for col in COLUMN_MAP['수령자연락처']:
        if col in df.columns:
            df[col] = df[col].apply(format_phone_number)
    
    return df

def classify_day_and_dawn(file_path, address_col=None, progress_callback=None):
    """당일&새벽 한방양식 일괄구분"""
    
    df = pd.read_excel(file_path, dtype=str)
    api_key = load_api_key()
    day_zip_codes = load_zip_codes('day')
    dawn_zip_codes = load_zip_codes('dawn')
    
    logger.info(f"당일배송 우편번호: {len(day_zip_codes)}개")
    logger.info(f"새벽배송 우편번호: {len(dawn_zip_codes)}개")
    
    # 주소 컬럼 찾기
    if not address_col:
        for col in ADDRESS_CANDIDATES:
            if col in df.columns:
                address_col = col
                break
    
    if not address_col:
        raise ValueError("주소 컬럼을 찾을 수 없습니다.")
    
    result_rows = []
    total_rows = len(df)
    processed = 0
    
    for idx, row in df.iterrows():
        if progress_callback:
            progress = (processed / total_rows) * 100
            progress_callback(progress)
        
        address = row[address_col]
        zipcode = get_zipcode_from_address(address, api_key)
        
        qty = 1
        qty_val = find_column(row, '수량')
        if qty_val:
            try:
                qty = int(qty_val)
            except:
                qty = 1
        
        if zipcode:
            zipcode_str = str(zipcode).zfill(5)
            for _ in range(qty):
                new_row = row.copy()
                if zipcode_str in day_zip_codes:
                    new_row['집하지점코드'] = ''  # 당일배송
                elif zipcode_str in dawn_zip_codes:
                    new_row['집하지점코드'] = '108'  # 새벽배송
                else:
                    continue  # 당일/새벽이 아니면 제외
                
                # 수량 컬럼 1로 설정
                for col in df.columns:
                    if '수량' in col:
                        new_row[col] = '1'
                
                result_rows.append(new_row)
        
        processed += 1
    
    if progress_callback:
        progress_callback(100)
    
    if not result_rows:
        return None
    
    result_df = pd.DataFrame(result_rows)
    
    # 주소 합치기 및 상세주소 처리 (당일배송과 동일한 로직 적용)
    # 1. 상세주소 컬럼 찾기
    detail_col = None
    for col in COLUMN_MAP['수령인상세주소']:
        if col in result_df.columns:
            detail_col = col
            break
    
    # 2. 주소 합치기 (당일배송에서 사용하는 방식과 동일)
    if detail_col and detail_col in result_df.columns and address_col in result_df.columns:
        def make_full_addr(row):
            base = str(row[address_col]) if pd.notna(row[address_col]) else ''
            detail = str(row[detail_col]) if pd.notna(row[detail_col]) and str(row[detail_col]).strip() else ''
            base = base.rstrip(', ').strip()
            # '0 / 0' 처리 추가
            if detail and detail.lower() != 'nan' and detail != '0 / 0':
                return f"{base} {detail}".strip()
            else:
                return base
        result_df['최종주소'] = result_df.apply(make_full_addr, axis=1)
        
        # 3. 최종주소를 받는분주소에 복사하고 상세주소 컬럼들 삭제
        result_df[address_col] = result_df['최종주소'].apply(clean_address_for_output)
        
        # 상세주소 컬럼들 삭제
        detail_cols_to_remove = []
        for col in COLUMN_MAP['수령인상세주소']:
            if col in result_df.columns:
                detail_cols_to_remove.append(col)
        
        # 추가 패턴으로 상세주소 컬럼 찾기
        for col in result_df.columns:
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in ['상세주소', '상세 주소', 'detail']):
                if col not in detail_cols_to_remove:
                    detail_cols_to_remove.append(col)
        
        if detail_cols_to_remove:
            result_df = result_df.drop(columns=detail_cols_to_remove, errors='ignore')
        
        # 최종주소 컬럼 제거
        result_df = result_df.drop(columns=['최종주소'])
    else:
        # 상세주소가 없으면 기본 주소만 정리
        if address_col in result_df.columns:
            result_df[address_col] = result_df[address_col].apply(clean_address_for_output)
    
    # 전화번호 포맷팅
    for col in COLUMN_MAP['수령인연락처1']:
        if col in result_df.columns:
            result_df[col] = result_df[col].apply(format_phone_number)
    
    # 결과 저장 (사용자 데이터 디렉토리에 저장)
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "user")
    os.makedirs(output_dir, exist_ok=True)
    output_path = get_incremented_filename("당일새벽배송_주문리스트", "xlsx", output_dir)
    save_with_yellow_header(result_df, output_path)
    
    return output_path

def classify_all_delivery_types(file_path, address_col=None, progress_callback=None):
    """당일&새벽&택배 한방양식 일괄구분"""
    
    df = pd.read_excel(file_path, dtype=str)
    api_key = load_api_key()
    day_zip_codes = load_zip_codes('day')
    dawn_zip_codes = load_zip_codes('dawn')
    
    logger.info(f"당일배송 우편번호: {len(day_zip_codes)}개")
    logger.info(f"새벽배송 우편번호: {len(dawn_zip_codes)}개")
    
    # 주소 컬럼 찾기
    if not address_col:
        for col in ADDRESS_CANDIDATES:
            if col in df.columns:
                address_col = col
                break
    
    if not address_col:
        raise ValueError("주소 컬럼을 찾을 수 없습니다.")
    
    result_rows = []
    address_error_rows = []
    total_rows = len(df)
    processed = 0
    
    for idx, row in df.iterrows():
        if progress_callback:
            progress = (processed / total_rows) * 100
            progress_callback(progress)
        
        address = row[address_col]
        zipcode = get_zipcode_from_address(address, api_key)
        
        qty = 1
        qty_val = find_column(row, '수량')
        if qty_val:
            try:
                qty = int(qty_val)
            except:
                qty = 1
        
        if not zipcode:
            address_error_rows.append(row)
            processed += 1
            continue
        
        zipcode_str = str(zipcode).zfill(5)
        
        for _ in range(qty):
            new_row = row.copy()
            if zipcode_str in day_zip_codes:
                new_row['집하지점코드'] = ''  # 당일배송
            elif zipcode_str in dawn_zip_codes:
                new_row['집하지점코드'] = '108'  # 새벽배송
            else:
                new_row['집하지점코드'] = '105'  # 택배
            
            # 수량 컬럼 1로 설정
            for col in df.columns:
                if '수량' in col:
                    new_row[col] = '1'
            
            result_rows.append(new_row)
        
        processed += 1
    
    if progress_callback:
        progress_callback(100)
    
    # 결과 데이터프레임 생성
    if '집하지점코드' not in df.columns:
        df_columns = df.columns.tolist() + ['집하지점코드']
    else:
        df_columns = df.columns.tolist()
    
    result_df = pd.DataFrame(result_rows, columns=df_columns)
    
    # 주소 합치기 및 상세주소 처리 (당일배송과 동일한 로직 적용)
    # 1. 상세주소 컬럼 찾기
    detail_col = None
    for col in COLUMN_MAP['수령인상세주소']:
        if col in result_df.columns:
            detail_col = col
            break
    
    # 2. 주소 합치기 (당일배송에서 사용하는 방식과 동일)
    if detail_col and detail_col in result_df.columns and address_col in result_df.columns:
        def make_full_addr(row):
            base = str(row[address_col]) if pd.notna(row[address_col]) else ''
            detail = str(row[detail_col]) if pd.notna(row[detail_col]) and str(row[detail_col]).strip() else ''
            base = base.rstrip(', ').strip()
            # '0 / 0' 처리 추가
            if detail and detail.lower() != 'nan' and detail != '0 / 0':
                return f"{base} {detail}".strip()
            else:
                return base
        result_df['최종주소'] = result_df.apply(make_full_addr, axis=1)
        
        # 3. 최종주소를 받는분주소에 복사하고 상세주소 컬럼들 삭제
        result_df[address_col] = result_df['최종주소'].apply(clean_address_for_output)
        
        # 상세주소 컬럼들 삭제
        detail_cols_to_remove = []
        for col in COLUMN_MAP['수령인상세주소']:
            if col in result_df.columns:
                detail_cols_to_remove.append(col)
        
        # 추가 패턴으로 상세주소 컬럼 찾기
        for col in result_df.columns:
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in ['상세주소', '상세 주소', 'detail']):
                if col not in detail_cols_to_remove:
                    detail_cols_to_remove.append(col)
        
        if detail_cols_to_remove:
            result_df = result_df.drop(columns=detail_cols_to_remove, errors='ignore')
        
        # 최종주소 컬럼 제거
        result_df = result_df.drop(columns=['최종주소'])
    else:
        # 상세주소가 없으면 기본 주소만 정리
        if address_col in result_df.columns:
            result_df[address_col] = result_df[address_col].apply(clean_address_for_output)
    
    # 결과 저장 (사용자 데이터 디렉토리에 저장)
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "user")
    os.makedirs(output_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(file_path))[0] + '_당일새벽택배한방양식'
    output_path = get_incremented_filename(base, 'xlsx', output_dir)
    save_with_yellow_header(result_df, output_path)
    
    # 주소오류 저장 (당일배송과 동일한 처리 적용)
    if address_error_rows:
        df_addrerr = process_unclassified_data(address_error_rows, address_col, detail_col)
        error_path = get_incremented_filename("주소오류_미분류", "xlsx", output_dir)
        save_with_yellow_header(df_addrerr, error_path)
    
    return output_path