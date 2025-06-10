import os
import pandas as pd
from datetime import datetime
from utils.file_utils import read_excel_file
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# 로깅 설정
logger = logging.getLogger(__name__)

def apply_dawn_option(
    template_path: str,
    request_type: str,
    msg_type: str,
    cycle: int
) -> str:
    """
    1) 새벽배송 양식 엑셀 읽기
    2) '요청유형', '문자전송유형', '차수' 컬럼에 값 채워 넣기
    3) 결과 엑셀 저장 후 경로 반환
    """
    try:
        # 입력 파일 읽기
        logger.info(f"새벽옵션 템플릿 파일 읽기: {template_path}")
        df = read_excel_file(template_path)
        
        # 요청유형, 문자전송유형, 차수 설정
        if "요청유형" in df.columns:
            df["요청유형"] = request_type
        elif "배송 유형" in df.columns:
            df["배송 유형"] = request_type
            
        if "문자전송유형" in df.columns:
            df["문자전송유형"] = msg_type
        elif "배송 문자 전송 시점" in df.columns:
            df["배송 문자 전송 시점"] = msg_type
            
        # 차수 추가 (숫자차)
        chasu_str = ""
        if cycle > 0:
            chasu_str = f"{cycle}차"
            
        if "차수" in df.columns:
            df["차수"] = chasu_str
        
        # 배송 받을 장소 상세 기본값 설정
        if "배송 받을 장소 상세" in df.columns:
            df["배송 받을 장소 상세"] = df["배송 받을 장소 상세"].fillna("문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송")
            # 빈 문자열인 경우에도 기본값 설정
            df.loc[df["배송 받을 장소 상세"] == "", "배송 받을 장소 상세"] = "문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송"
            # NaN 값이나 None 값도 처리
            df.loc[df["배송 받을 장소 상세"].isna(), "배송 받을 장소 상세"] = "문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송"
        elif "배송 받을 장소 상세*" in df.columns:
            df["배송 받을 장소 상세*"] = df["배송 받을 장소 상세*"].fillna("문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송")
            # 빈 문자열인 경우에도 기본값 설정
            df.loc[df["배송 받을 장소 상세*"] == "", "배송 받을 장소 상세*"] = "문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송"
            # NaN 값이나 None 값도 처리
            df.loc[df["배송 받을 장소 상세*"].isna(), "배송 받을 장소 상세*"] = "문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송"
        
        # 배송 받을 장소 설정
        if "배송 받을 장소" in df.columns:
            df["배송 받을 장소"] = df["배송 받을 장소"].fillna("문 앞")
            df.loc[df["배송 받을 장소"] == "", "배송 받을 장소"] = "문 앞"
        
        # 상품명 포맷팅
        if "상품명*" in df.columns:
            df["상품명*"] = df.apply(lambda row: format_product_name(row, chasu_str), axis=1)
        elif "상품명" in df.columns:
            df["상품명"] = df.apply(lambda row: format_product_name(row, chasu_str), axis=1)
        
        # 배송요청일에서 "00:00:00" 부분 제거
        for col in df.columns:
            if "배송요청일" in col or "배송 요청일" in col:
                if df[col].dtype == 'datetime64[ns]':
                    df[col] = df[col].dt.strftime('%Y-%m-%d')
                else:
                    # 문자열인 경우 "00:00:00" 부분 제거
                    df[col] = df[col].astype(str).str.replace(r' 00:00:00(\.\d+)?$', '', regex=True)
        
        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "user")
        os.makedirs(output_dir, exist_ok=True)
        
        # GUI 프로그램과 동일한 파일명 사용
        filename = f"새벽배송_옵션추가완료_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, filename)
        
        logger.info(f"새벽옵션 추가 완료: 행 수={len(df)}, 열 수={len(df.columns)}")
        df.to_excel(output_path, index=False, engine="openpyxl")
        
        # 파스텔톤 배경색 적용 (주소 중복 시)
        apply_pastel_tone_for_duplicates(output_path)
        
        logger.info(f"파일 저장 완료: {output_path}")
        
        # 웹 경로로 변환 (API를 통해 다운로드 가능하도록)
        web_path = f"/api/data/download/{filename}"
        return web_path
        
    except Exception as e:
        logger.error(f"새벽옵션 추가 중 오류 발생: {str(e)}")
        raise ValueError(f"새벽옵션 추가 중 오류가 발생했습니다: {str(e)}")

def format_product_name(row, chasu_str):
    """상품명을 '주문 : [마스킹된주문자명] + 차수 + 회사명 + 상품명' 형식으로 변환"""
    # 주문자명/구매자명 찾기
    buyer_name = ""
    for col in ["주문자*", "주문자명", "주문자", "구매자명", "구매자"]:
        if col in row and pd.notna(row[col]):
            buyer_name = str(row[col])
            break
    
    # 주문자명 마스킹 처리
    masked_buyer = ""
    if buyer_name:
        if len(buyer_name) == 2:  # 두 글자 이름
            masked_buyer = buyer_name[0] + '*'
        elif len(buyer_name) >= 3:  # 세 글자 이상 이름
            masked_buyer = buyer_name[0] + '*' * (len(buyer_name) - 2) + buyer_name[-1]
        else:  # 한 글자 이름
            masked_buyer = buyer_name
    
    # 회사명 추가
    company = ""
    
    # 거래처 주문번호 관련 컬럼에서 남촌상회 확인
    order_columns = ["거래처 주문번호", "거래처주문번호", "주문번호", "거래처 번호", "거래처번호"]
    for col in order_columns:
        if col in row and pd.notna(row[col]):
            value = str(row[col]).lower()
            if "남촌" in value or "남촌상회" in value or "남촌과일" in value:
                company = " 남촌상회"
                logger.info(f"거래처 주문번호에서 남촌상회 발견: {value}")
                break
    
    # 거래처 주문번호에서 회사명을 찾지 못한 경우 다른 컬럼에서 검색
    if not company:
        # 기타 컬럼 검색 (주문자 정보 등)
        addr_columns = [
            "송하인주소", "송하인 주소", "발송인주소", "발신인주소", 
            "주문자 주소", "주문자주소", "구매자 주소", "구매자주소",
            "송하인명", "송하인이름", "발송인명", "발신인명",
            "주문자명", "주문자", "구매자명", "구매자",
            "상품명", "상품명*"
        ]
        
        for col in addr_columns:
            if col in row and pd.notna(row[col]):
                value = str(row[col]).lower()
                
                if "총각네" in value or "자연의모든것" in value or "자연의 모든것" in value:
                    company = " 총각네"
                    break
                elif "이낙근" in value:
                    company = " 이낙근"
                    break
                elif "국앤찬" in value:
                    company = " 국앤찬"
                    break
                elif "맛사랑" in value:
                    company = " 맛사랑"
                    break
                elif "새농" in value:
                    company = " 새농"
                    break
                elif "당일장터" in value:
                    company = " 당일장터"
                    break
                elif "남촌" in value or "남촌상회" in value or "남촌과일" in value:
                    company = " 남촌상회"
                    break
    
    # 상품명 가져오기
    product = ""
    for col in ["상품명*", "상품명"]:
        if col in row and pd.notna(row[col]):
            product = str(row[col])
            break
    
    # 최종 상품명 형식 생성
    prefix = "주문 : " + masked_buyer
    result = f"{prefix}{company}{' ' + chasu_str if chasu_str else ''} {product}".strip()
    
    # 디버깅용 로그 (거래처 주문번호 관련)
    for col in order_columns:
        if col in row and pd.notna(row[col]):
            logger.info(f"거래처 주문번호 확인: {col}={row[col]}, 추출된 회사명: {company}")
    
    return result

def apply_pastel_tone_for_duplicates(file_path):
    """주소 중복에 파스텔톤 배경색 적용"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 주소 컬럼 찾기 (받는분주소를 우선적으로 찾기)
        address_col = None
        address_headers = ['받는분주소', '수령인주소', '수취인주소', '수령자주소', '배송지주소', '배송 주소', '주소', '배송지']
        
        for i, cell in enumerate(ws[1], 1):
            value = str(cell.value or '').strip()
            # 상세가 포함되지 않은 주소 컬럼 찾기
            if any(header in value for header in address_headers) and '상세' not in value and '송하인' not in value:
                address_col = i
                logger.info(f"주소 컬럼 발견: {value} (열 {i})")
                break
        
        if not address_col:
            # 다시 한번 더 넓은 범위로 검색 (송하인주소 제외)
            for i, cell in enumerate(ws[1], 1):
                value = str(cell.value or '').strip()
                if any(header in value for header in address_headers) and '송하인' not in value:
                    address_col = i
                    logger.info(f"주소 컬럼 발견(2차 검색): {value} (열 {i})")
                    break
        
        if address_col:
            logger.info(f"주소 컬럼 인덱스: {address_col}, 값: {ws.cell(row=1, column=address_col).value}")
            
            # 주소별 행 매핑
            addr_to_rows = {}
            
            # 모든 주소 수집 (2행부터 시작)
            for row in range(2, ws.max_row + 1):
                addr = str(ws.cell(row=row, column=address_col).value or '').strip()
                if not addr or addr.lower() in ('nan', 'none', '0'):
                    continue
                
                # 주소 정규화
                norm_addr = normalize_address(addr)
                if norm_addr == '':
                    continue
                
                if norm_addr not in addr_to_rows:
                    addr_to_rows[norm_addr] = []
                addr_to_rows[norm_addr].append(row)
            
            logger.info(f"중복 주소 검사: 총 {len(addr_to_rows)} 개의 고유 주소 발견")
            
            # 중복 주소 찾기 (2개 이상 같은 주소가 있는 경우)
            duplicate_addresses = {addr: rows for addr, rows in addr_to_rows.items() if len(rows) >= 2}
            logger.info(f"중복 주소 수: {len(duplicate_addresses)}")
            
            if duplicate_addresses:
                colors = [
                    "FFFF99", "FFEBEE", "E3F2FD", "E8F5E9", "FFFDE7", "F3E5F5", "FBE9E7",
                    "E0F2F1", "FFF3E0", "F9FBE7", "EDE7F6", "F1F8E9", "FCE4EC"
                ]
                
                # 각 중복 주소 그룹에 색상 할당
                color_idx = 0
                
                # 헤더에 노란색 적용
                yellow_fill = PatternFill(
                    start_color='FFFF00',
                    end_color='FFFF00',
                    fill_type='solid'
                )
                for cell in ws[1]:
                    cell.fill = yellow_fill
                
                # 중복 주소별로 색상 적용
                for addr, rows in duplicate_addresses.items():
                    fill = PatternFill(
                        start_color=colors[color_idx % len(colors)],
                        end_color=colors[color_idx % len(colors)],
                        fill_type='solid'
                    )
                    color_idx += 1
                    
                    logger.info(f"중복 주소 '{addr}' 행 수: {len(rows)}, 색상: {colors[(color_idx-1) % len(colors)]}")
                    
                    # 해당 행의 모든 셀에 배경색 적용
                    for row_idx in rows:
                        for cell in ws[row_idx]:
                            cell.fill = fill
            else:
                logger.info("중복 주소가 없습니다.")
        else:
            logger.warning("주소 컬럼을 찾을 수 없습니다.")
        
        # 변경사항 저장
        wb.save(file_path)
        logger.info(f"파스텔톤 적용 완료 및 파일 저장: {file_path}")
        
    except Exception as e:
        logger.error(f"파스텔톤 적용 중 오류 발생: {str(e)}")
        # 오류가 있어도 진행

def normalize_address(addr):
    """주소 정규화 (파스텔톤 적용을 위한 주소 비교용)"""
    import re
    if not addr:
        return ''
    
    s = str(addr).strip()
    if s == '' or s.lower() == 'nan' or s.lower() == 'none' or s == '0' or s == '0 / 0':
        return ''
    
    # 모든 공백, 줄바꿈, 탭, 유니코드 공백 제거
    s = re.sub(r'[\s\r\n\t\u200b\u3000]', '', s)
    
    # 괄호 및 괄호 안 내용 제거
    s = re.sub(r'\(.*?\)', '', s)
    
    # 특정 패턴 통일 (예: 1동 1호 -> 1동1호)
    s = re.sub(r'(\d+)[동](\d+)[호]', r'\1동\2호', s)
    
    # 동,호,층,세대,실,번지 등 단위 통일
    s = re.sub(r'([0-9]+)[ ]*동', r'\1동', s)
    s = re.sub(r'([0-9]+)[ ]*호', r'\1호', s)
    s = re.sub(r'([0-9]+)[ ]*층', r'\1층', s)
    s = re.sub(r'([0-9]+)[ ]*세대', r'\1세대', s)
    s = re.sub(r'([0-9]+)[ ]*실', r'\1실', s)
    s = re.sub(r'([0-9]+)[ ]*번지', r'\1번지', s)
    
    # 특수문자 제거 (한글, 영문, 숫자만 남김)
    s = re.sub(r'[^가-힣a-zA-Z0-9]', '', s)
    
    # 아파트, 빌라, 맨션 등 동일한 의미의 단어 통일
    s = s.replace('아파트', '아파트')
    s = s.replace('빌라', '빌라')
    s = s.replace('맨션', '맨션')
    s = s.replace('오피스텔', '오피스텔')
    
    return s.lower()  # 소문자로 변환하여 반환 