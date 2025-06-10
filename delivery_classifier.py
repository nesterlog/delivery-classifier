import pandas as pd
import requests
import json
import os
import re
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter.font as tkfont
from PIL import Image, ImageTk
import difflib
import win32com.client
import time
import random
import queue
import threading
import gc
import psutil
import warnings

SAME_DAY_FILE = 'data/당일_우편번호.csv'
DAWN_FILE = 'data/새벽_우편번호.csv'
API_KEY_FILE = 'data/api_key.txt'
DEFAULT_KAKAO_API_KEY = 'c3207609f232ab9972310f876e22e233'

ADDRESS_CANDIDATES = [
    '받는분주소', '받는분 주소', '수취인주소', '수취인 주소', '주소', '배송지주소', '배송지 주소', '수령인 주소(전체)', '수령인 주소', '수령인주소'
]

COLUMN_MAP = {
    '수령자': ['수령자', '수령인', '받는분', '수취인', '수취인명', '수령인명'],
    '수령자도로명주소': ['수령자 도로명 주소', '수령인 주소*', '수령인 주소(전체)', '받는분주소', '주소', '수령인 도로명 주소'],
    '수령인주소': ['수령자 도로명 주소', '수령인 주소*', '수령인 주소(전체)', '받는분주소', '주소', '수령인 도로명 주소'],
    '수령자상세주소': ['수령자 상세주소', '수령인 상세주소*', '상세주소', '상세 주소', '수령인상세주소', '받는분상세주소', '수령인 상세 주소', '수령자 상세 주소'],
    '수령인상세주소': ['수령자 상세주소', '수령인 상세주소*', '상세주소', '상세 주소', '수령인상세주소', '받는분상세주소', '수령인 상세 주소', '수령자 상세 주소'],
    '수령자연락처': ['수령자 연락처', '수령인 핸드폰*', '수령인 연락처', '수령인 휴대전화', '수취인연락처1'],
    '수령인연락처1': ['수령자 연락처', '수령인 핸드폰*', '수령인 연락처', '수령인 휴대전화', '수취인연락처1'],
    '수령인연락처2': ['수령자 연락처2', '수령인 연락처2', '수령인 휴대전화2', '수취인연락처2'],
    '구매자연락처': ['구매자 연락처', '주문자 연락처', '구매자연락처', '주문자연락처', '구매자핸드폰', '주문자핸드폰'],
    '상품명': ['상품명*', '상품명(총각네 쇼핑몰)', '상품명'],
    '거래처주문코드': ['거래처 주문번호', '거래처주문코드', '주문번호', '상품주문번호'],
    '주문자': ['주문자*', '주문자명', '주문자', '구매자명', '구매자', '보내는분'],
    '수량': ['수량'],
    '배송메시지': ['배송메시지', '비고2(배송메시지)', '출입방법', '출입 방법'],
    '배송받을장소': ['배송 받을 장소', '비고'],
    '배송받을장소상세': ['배송 받을 장소 상세'],
    '배송유형': ['배송 유형', '요청유형*'],
    '배송문자전송시점': ['배송 문자 전송 시점', '배송문자유형'],
    '주문자주소': ['주문자 주소', '송하인주소', '구매자 주소', '주문자주소', '보내는분 주소', '수취인 주소', '받는분 주소'],
}

VERSION = "2.0"

def preprocess_address(address):
    import re
    # 기존 전처리 로직 (괄호, 동/호/아파트/빌라/층 등 제거)
    address = re.sub(r'\(.*?\)', '', address)
    address = re.sub(r'\d+동|\d+호|\d+층|아파트|빌라|APT|apt|센트레빌|골드클래스|현대빌라|리버파크', '', address, flags=re.IGNORECASE)
    address = re.sub(r'\s+', ' ', address).strip()
    return address

def clean_address(address):
    # 괄호 및 괄호 안 내용 제거
    address = re.sub(r'\(.*?\)', '', address)
    # 동/호수 등 숫자-숫자 패턴 제거
    address = re.sub(r'\d+[\-~]\d+호?', '', address)
    # '^^' 제거
    address = address.replace('^^', '')
    # 불필요한 공백 정리
    address = re.sub(r'\s+', ' ', address).strip()
    return address

def clean_address_for_output(address):
    # '@' → '아파트'
    address = address.replace('@', '아파트')
    # 별모양 특수문자 → '*'
    for star in ['★', '☆', '✪', '✯', '✩', '✬', '✭', '✮', '✰', '✶', '✷', '✸', '✹', '✺', '✻', '✼', '✽', '✾', '✿', '❀', '❁', '❂', '❃', '❄', '❅', '❆', '❇', '❈', '❉', '❊', '❋']:
        address = address.replace(star, '*')
    # 로마숫자 → 숫자
    roman_map = {'Ⅰ': '1', 'Ⅱ': '2', 'Ⅲ': '3', 'Ⅳ': '4', 'Ⅴ': '5', 'Ⅵ': '6', 'Ⅶ': '7', 'Ⅷ': '8', 'Ⅸ': '9', 'Ⅹ': '10'}
    for k, v in roman_map.items():
        address = address.replace(k, v)
    # '--', 'ㅡ', '~' → '-', '&' → '앤'
    address = address.replace('--', '-').replace('ㅡ', '-').replace('~', '-').replace('&', '앤')
    return address

def find_column(row, key):
    import re
    candidates = COLUMN_MAP.get(key, [])
    row_keys = [str(k).strip() for k in row.keys()]
    def normalize(s):
        return re.sub(r'[^가-힣a-zA-Z0-9]', '', str(s)).lower()
    # 1. 완전일치
    for col in candidates:
        for rk in row_keys:
            if col == rk:
                return row.get(rk, '')
    # 2. normalize 일치
    for col in candidates:
        col_base = normalize(col)
        for rk in row_keys:
            rk_base = normalize(rk)
            if col_base == rk_base:
                return row.get(rk, '')
    # 3. 부분일치(가장 짧은 컬럼 우선)
    matches = []
    for col in candidates:
        col_base = normalize(col)
        for rk in row_keys:
            rk_base = normalize(rk)
            if col_base in rk_base or rk_base in col_base:
                matches.append((len(rk_base), rk, row.get(rk, '')))
    if matches:
        matches.sort()
        return matches[0][2]
    # 4. fallback: key in rk, rk in key (normalize)
    key_norm = normalize(key)
    fallback = []
    for rk in row_keys:
        rk_base = normalize(rk)
        if key_norm in rk_base or rk_base in key_norm:
            fallback.append((len(rk_base), rk, row.get(rk, '')))
    if fallback:
        fallback.sort()
        return fallback[0][2]
    return ''

def replace_company_name(address):
    if not isinstance(address, str):
        return address
    return address.replace('(주)자연의 모든것', '총각네')

def has_jonggakne_company(addr):
    if not isinstance(addr, str):
        return False
    # (주), ㈜, 공백 등 다양한 변형 매칭
    return bool(re.search(r'(\(주\)|㈜)\s*자연의\s*모든것', addr))

def has_nakgeun_text(addr):
    if not isinstance(addr, str):
        return False
    return '이낙근' in addr

def has_gooknchan_text(addr):
    if not isinstance(addr, str):
        return False
    return '국앤찬' in addr

def format_phone_number(phone):
    import re
    if pd.isna(phone):  # NaN 체크 추가
        return ''
    if not isinstance(phone, str):
        phone = str(phone)
    phone = phone.strip().replace('-', '')
    # 050x-xxxx-xxxx (12자리)
    if re.match(r'^050[0-9]', phone) and len(phone) == 12:
        return f"{phone[:4]}-{phone[4:8]}-{phone[8:]}"
    # 02-xxx-xxxx or 02-xxxx-xxxx
    if phone.startswith('02') and len(phone) in (9, 10):
        if len(phone) == 9:
            return f"{phone[:2]}-{phone[2:5]}-{phone[5:]}"
        else:
            return f"{phone[:2]}-{phone[2:6]}-{phone[6:]}"
    # 3자리 지역번호(031, 051 등)
    elif re.match(r'^0[3-6][0-9]', phone) and len(phone) in (10, 11):
        if len(phone) == 10:
            return f"{phone[:3]}-{phone[3:6]}-{phone[6:]}"
        else:
            return f"{phone[:3]}-{phone[3:7]}-{phone[7:]}"
    # 휴대폰(010, 011, 016, 017, 018, 019)
    elif re.match(r'^01[016789]', phone) and len(phone) == 11:
        return f"{phone[:3]}-{phone[3:7]}-{phone[7:]}"
    elif re.match(r'^01[016789]', phone) and len(phone) == 10:
        return f"{phone[:3]}-{phone[3:6]}-{phone[6:]}"
    # 1588, 1577, 16xx, 18xx 등 4-4
    elif re.match(r'^(15|16|18)[0-9]{2}', phone) and len(phone) == 8:
        return f"{phone[:4]}-{phone[4:]}"
    elif re.match(r'^070', phone) and len(phone) == 11:
        return f"{phone[:3]}-{phone[3:7]}-{phone[7:]}"
    # 이미 -가 있거나 기타: 그대로 반환
    return phone

def convert_xls_to_xlsx(xls_path):
    """Convert .xls file to .xlsx format using a more robust method.
    This function handles the conversion while preserving formatting and data types.
    """
    import win32com.client
    import os
    
    # Get absolute path
    abs_path = os.path.abspath(xls_path)
    xlsx_path = abs_path + '.xlsx'
    
    try:
        # Try using win32com (most reliable method)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            wb = excel.Workbooks.Open(abs_path)
            wb.SaveAs(xlsx_path, FileFormat=51)  # 51 is for .xlsx format
            wb.Close()
            return xlsx_path
        finally:
            excel.Quit()
    except Exception as e:
        # Fallback to pandas if win32com fails
        try:
            df = pd.read_excel(xls_path, engine='xlrd')
            df.to_excel(xlsx_path, index=False, engine='openpyxl')
            return xlsx_path
        except Exception as inner_e:
            raise Exception(f"Failed to convert file: {str(e)}. Fallback also failed: {str(inner_e)}")

def read_excel_file(file_path, **kwargs):
    """Read Excel file with automatic format handling and memory management."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
    if not os.access(file_path, os.R_OK):
        raise PermissionError(f"파일 읽기 권한이 없습니다: {file_path}")
    ext = os.path.splitext(file_path)[1].lower()
    temp_file = None
    try:
        if ext == '.xls':
            temp_file = convert_xls_to_xlsx(file_path)
            file_path = temp_file
        df = pd.read_excel(file_path, engine='openpyxl', **kwargs)
        return df
    except Exception as e:
        raise Exception(f"Excel 파일 읽기 실패: {str(e)}")
    finally:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass

def read_with_template(input_file, template_columns):
    try:
        df = read_excel_file(input_file, dtype=str, header=None)
        first_row = df.iloc[0]
        if first_row.isnull().all() or first_row.astype(str).str.match(r'^(Unnamed|\d+)?$').all():
            df.columns = template_columns
            df = df.iloc[1:]
        else:
            df.columns = [str(x).strip() if pd.notna(x) else f'Column_{i}' for i, x in enumerate(first_row)]
            df = df.iloc[1:]
        return df.reset_index(drop=True)
    except Exception as e:
        print(f"템플릿 읽기 오류: {str(e)}")
        raise

class DeliveryClassifier:
    def __init__(self, api_key=None):
        # api_key.txt가 없으면 기본값을 파일로 저장
        if not os.path.exists(API_KEY_FILE):
            with open(API_KEY_FILE, 'w', encoding='utf-8') as f:
                f.write(DEFAULT_KAKAO_API_KEY)
        self.api_key = api_key or self.load_api_key() or DEFAULT_KAKAO_API_KEY
        self.same_day_zipcodes = set()
        self.dawn_zipcodes = set()
        self.load_zipcode_lists()
        self._zipcode_cache = {}  # 주소→우편번호 캐시 추가

    def load_api_key(self):
        if os.path.exists(API_KEY_FILE):
            with open(API_KEY_FILE, 'r', encoding='utf-8') as f:
                return f.read().strip()
        return ''

    def save_api_key(self, api_key):
        with open(API_KEY_FILE, 'w', encoding='utf-8') as f:
            f.write(api_key.strip())
        self.api_key = api_key.strip()

    def load_zipcode_lists(self):
        if os.path.exists(SAME_DAY_FILE):
            same_day_df = pd.read_csv(SAME_DAY_FILE, header=None)
            self.same_day_zipcodes = set(same_day_df[0].astype(str))
        if os.path.exists(DAWN_FILE):
            dawn_df = pd.read_csv(DAWN_FILE, header=None)
            self.dawn_zipcodes = set(dawn_df[0].astype(str))

    def update_zipcode_file(self, filetype, new_file):
        if filetype == 'same_day':
            os.replace(new_file, SAME_DAY_FILE)
        elif filetype == 'dawn':
            os.replace(new_file, DAWN_FILE)
        self.load_zipcode_lists()

    def extract_road_address_only(self, address: str) -> str:
        # 도로명+건물번호만 추출 (예: '별내5로 189-130')
        import re
        m = re.search(r'([가-힣A-Za-z0-9\s]+(로|길))\s*\d+[\-\d]*', address)
        if m:
            return m.group(0).strip()
        return address.split()[0] if address else address

    def get_zipcode_from_address(self, address: str) -> str:
        # address가 문자열이 아니면 빈 문자열로 처리 (방어코드)
        if not address or not isinstance(address, str):
            return ""
        # 캐시 먼저 확인
        if address in self._zipcode_cache:
            return self._zipcode_cache[address]
        import re
        tried = set()
        addr_try_list = []
        seen = set()
        # 1. 도로명+건물번호
        m = re.search(r'([가-힣A-Za-z0-9\s]+(로|길))\s*\d+[\-\d]*', address)
        if m:
            road_addr = m.group(0).strip()
            if road_addr and road_addr not in seen:
                addr_try_list.append(road_addr)
                seen.add(road_addr)
        # 2. 지번주소(동+번지)
        m2 = re.search(r'([가-힣A-Za-z0-9\s]+동)\s*\d+[\-\d]*', address)
        if m2:
            jibun_addr = m2.group(0).strip()
            if jibun_addr and jibun_addr not in seen:
                addr_try_list.append(jibun_addr)
                seen.add(jibun_addr)
        # 3. 원본 주소
        if address and address not in seen:
            addr_try_list.append(address)
            seen.add(address)
        # 4. 괄호제거
        bracketless = re.sub(r'\(.*?\)', '', address).strip()
        if bracketless and bracketless not in seen:
            addr_try_list.append(bracketless)
            seen.add(bracketless)
        # 5. 기타 전처리
        cleaned = clean_address(address)
        if cleaned and cleaned not in seen:
            addr_try_list.append(cleaned)
            seen.add(cleaned)
        cleaned_out = clean_address_for_output(address)
        if cleaned_out and cleaned_out not in seen:
            addr_try_list.append(cleaned_out)
            seen.add(cleaned_out)
        # 6. 마지막 시도: 동/로까지만 남기고 API에 질의
        road_only = re.sub(r'\s+\d+.*', '', address)
        if road_only and road_only not in seen:
            addr_try_list.append(road_only)
            seen.add(road_only)
        for addr_try in addr_try_list:
            if not addr_try or addr_try in tried:
                continue
            tried.add(addr_try)
            print(f"[DEBUG][ZIP] 카카오API 시도 주소: {addr_try}")
            zipcode = self._search_zipcode_kakao(addr_try)
            print(f"[DEBUG][ZIP] API 응답: {zipcode}")
            if zipcode:
                try:
                    zc_str = str(int(zipcode)).zfill(5)
                except Exception:
                    zc_str = zipcode
                self._zipcode_cache[address] = zc_str  # 캐시에 저장
                return zc_str
        self._zipcode_cache[address] = ""
        return ""

    def _search_zipcode_kakao(self, address: str) -> str:
        max_retries = 3
        retry_delay = 1  # 초 단위
        
        for attempt in range(max_retries):
            try:
                # API 호출 간 랜덤 지연 (0.1초 ~ 0.3초)
                time.sleep(0.1 + random.uniform(0, 0.2))
                
                url = "https://dapi.kakao.com/v2/local/search/address.json"
                headers = {"Authorization": f"KakaoAK {self.api_key}"}
                params = {"query": address}
                
                response = requests.get(url, headers=headers, params=params)
                print(f"[카카오API] 주소: {address}")
                print(f"[카카오API] status_code: {response.status_code}")
                print(f"[카카오API] 응답: {response.text}")
                
                if response.status_code == 200:
                    data = response.json()
                    if data['documents']:
                        doc = data['documents'][0]
                        if doc.get('road_address') and doc['road_address'].get('zone_no'):
                            return doc['road_address']['zone_no']
                        elif doc.get('address') and doc['address'].get('zone_no'):
                            return doc['address']['zone_no']
                    else:
                        print(f'카카오API 주소 미매칭: {address}')
                        return ""
                elif response.status_code == 429:  # Too Many Requests
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay * (attempt + 1))
                        continue
                else:
                    print(f'카카오API 호출 실패: {address}, status={response.status_code}')
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay)
                        continue
            except Exception as e:
                print(f"카카오API 오류: {address}: {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    continue
        return ""

    def save_with_yellow_header(self, df, output_path, header=None):
        def normalize_addr(addr):
            import re
            s = str(addr or '').strip()
            if s == '' or s.lower() == 'nan' or s.lower() == 'none' or s == '0':
                return ''
            # 모든 공백, 줄바꿈, 탭, 유니코드 공백 제거
            s = re.sub(r'[\s\r\n\t\u200b\u3000]', '', s)
            # 괄호 및 괄호 안 내용 제거
            s = re.sub(r'\(.*?\)', '', s)
            # 특수문자 제거(한글, 영문, 숫자만 남김)
            s = re.sub(r'[^가-힣a-zA-Z0-9]', '', s)
            return s.lower()
        df.to_excel(output_path, index=False, header=header if header else True)
        wb = load_workbook(output_path)
        ws = wb.active
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for cell in ws[1]:
            cell.fill = yellow_fill
        address_candidates = [
            '받는분주소', '받는분 주소', '수취인주소', '수취인 주소', '수령인 주소*', '수령인주소', '수령인 주소',
            '배송지주소', '배송지 주소', '수령자 도로명 주소*', '송하인주소', '주소', '수령자 주소',
            '수령인 도로명 주소', '수령자도로명주소', '수령인도로명주소', '받는분 도로명 주소', '받는분도로명주소',
            '배송지 도로명 주소', '배송지도로명주소'
        ]
        def normalize_colname(s):
            import re
            return re.sub(r'[\s\r\n\t\u200b\u3000]', '', str(s)).lower()

        # 헤더 강제 str 변환
        print('엑셀 헤더:', [str(cell.value or '').strip() for cell in ws[1]])
        norm_headers = [normalize_colname(str(cell.value or '').strip()) for cell in ws[1]]
        # 0. '받는분주소'가 있으면 무조건 address_col로 지정
        address_col = None
        for i, cell in enumerate(ws[1], 1):
            val = str(cell.value or '').strip()
            if val == '받는분주소':
                address_col = i
                break
        # 없으면 기존 address_col 탐지 로직 사용
        if address_col is None:
            norm_candidates = [normalize_colname(c) for c in address_candidates]
            norm_headers = [normalize_colname(str(cell.value or '').strip()) for cell in ws[1]]
            # 1. '주소'가 들어간 컬럼 모두 후보로
            addr_candidates = []
            for i, cell in enumerate(ws[1], 1):
                val = str(cell.value or '').strip()
                if '주소' in val:
                    addr_candidates.append(i)
            # 2. 후보가 여러 개면, 데이터가 가장 많이 채워진(빈값이 적은) 컬럼 우선
            if addr_candidates:
                max_filled = -1
                best_col = addr_candidates[0]
                for col in addr_candidates:
                    filled = 0
                    for row in range(2, ws.max_row+1):
                        v = ws.cell(row=row, column=col).value
                        if v not in [None, '', 'nan', 'NaN', '0']:
                            filled += 1
                    if filled > max_filled:
                        max_filled = filled
                        best_col = col
                address_col = best_col
            # 3. 그래도 없으면 완전일치/부분일치/포함/유사도 등 기존 로직
            if address_col is None:
                for i, norm_header in enumerate(norm_headers):
                    if norm_header in norm_candidates:
                        address_col = i + 1
                        break
                if address_col is None:
                    for i, norm_header in enumerate(norm_headers):
                        for cand in norm_candidates:
                            if cand in norm_header or norm_header in cand:
                                address_col = i + 1
                                break
                        if address_col:
                            break
                if address_col is None:
                    for i, cell in enumerate(ws[1], 1):
                        val = str(cell.value or '').strip()
                        if '도로명' in val or '주소' in val:
                            address_col = i
                            break
                if address_col is None:
                    try:
                        from difflib import SequenceMatcher
                        best_score = 0
                        best_idx = None
                        for i, norm_header in enumerate(norm_headers):
                            for cand in norm_candidates:
                                score = SequenceMatcher(None, norm_header, cand).ratio()
                                if score > best_score:
                                    best_score = score
                                    best_idx = i
                        if best_score >= 0.8 and best_idx is not None:
                            address_col = best_idx + 1
                    except:
                        pass
            # 4. 마지막 fallback: 컬럼이 하나라도 있으면 그걸로 강제 적용
            if address_col is None and len(ws[1]) > 0:
                address_col = 1
        # 디버깅 출력
        print('엑셀 헤더:', [str(cell.value or '').strip() for cell in ws[1]])
        if address_col is not None:
            addr_raws = [str(ws.cell(row=r, column=address_col).value or '').strip() for r in range(2, ws.max_row+1)]
            addr_norms = [normalize_addr(ws.cell(row=r, column=address_col).value) for r in range(2, ws.max_row+1)]
            print('주소 원본:', addr_raws)
            print('주소 normalize:', addr_norms)
        # 이하 기존 파스텔톤 색상 적용 로직 동일 (addr_to_rows에 빈값은 무시)
        if address_col is not None:
            addr_to_rows = {}
            for row in range(2, ws.max_row + 1):
                addr = str(ws.cell(row=row, column=address_col).value or '').strip()
                norm_addr = normalize_addr(addr)
                if norm_addr == '':
                    continue
                if norm_addr not in addr_to_rows:
                    addr_to_rows[norm_addr] = []
                addr_to_rows[norm_addr].append(row)
            if len(addr_to_rows) >= 2:
                colors = [
                    "FFFF99", "FFEBEE", "E3F2FD", "E8F5E9", "FFFDE7", "F3E5F5", "FBE9E7",
                    "E0F2F1", "FFF3E0", "F9FBE7", "EDE7F6", "F1F8E9", "FCE4EC"
                ]
                color_map = {}
                color_idx = 0
                for addr, rows in addr_to_rows.items():
                    if len(rows) < 2:
                        continue
                    color_map[addr] = PatternFill(start_color=colors[color_idx % len(colors)], end_color=colors[color_idx % len(colors)], fill_type='solid')
                    color_idx += 1
                for addr, rows in addr_to_rows.items():
                    if addr not in color_map:
                        continue
                    fill = color_map[addr]
                    for row in rows:
                        for cell in ws[row]:
                            cell.fill = fill
            for cell in ws[1]:
                cell.fill = yellow_fill
        wb.save(output_path)
        # 저장 후 fill 적용 확인
        try:
            wb2 = load_workbook(output_path)
            ws2 = wb2.active
            print('색상 fill 확인:', [ws2.cell(row=r, column=address_col).fill.start_color.rgb for r in range(2, min(4, ws2.max_row+1))])
        except Exception as e:
            print('fill 확인 오류:', e)

    def get_incremented_filename(self, base, ext, output_dir):
        # base에 확장자가 이미 있으면 제거
        if base.endswith(f'.{ext}'):
            base = base[:-(len(ext)+1)]
        filename = f"{base}.{ext}"
        full_path = os.path.join(output_dir, filename)
        i = 1
        while os.path.exists(full_path):
            filename = f"{base}({i}).{ext}"
            full_path = os.path.join(output_dir, filename)
            i += 1
        return full_path

    def classify_all(self, input_file: str, address_column: str, chasu=None, dawn_type='배송대행', sms_type='즉시전송', progress_callback=None):
        df = read_excel_file(input_file, dtype=str)
        # 필수 컬럼 결측치 처리
        필수키 = ['구매자명', '구매자연락처1', '상품주문번호', '송하인주소', '수취인', '수취인연락처1', '받는분주소', '상품명', '수량', '수령인', '수령자']
        def normalize_colname(s):
            import re
            return re.sub(r'[\\s\\r\\n\\t\\u200b\\u3000]', '', str(s)).lower()
        실제필수컬럼 = set()
        norm_excel_cols = [normalize_colname(c) for c in df.columns]
        for key in 필수키:
            candidates = COLUMN_MAP.get(key, [key])
            for cand in candidates:
                norm_cand = normalize_colname(cand)
                for col, norm_col in zip(df.columns, norm_excel_cols):
                    if norm_col == norm_cand:
                        실제필수컬럼.add(col)
        실제필수컬럼 = list(실제필수컬럼)
        print('[디버그] 실제필수컬럼:', 실제필수컬럼)
        print('[디버그] df.columns:', list(df.columns))
        for col in 실제필수컬럼:
            try:
                print(f'[디버그] [{col}] 값들:', df[col].values)
            except Exception as e:
                print(f'[디버그] [{col}] 값 출력 오류:', e)
        if 실제필수컬럼:
            def is_missing(val):
                import pandas as pd, re
                if pd.isnull(val):
                    return True
                s = str(val).strip()
                s = re.sub(r'[\s\r\n\t\u200b\u3000]', '', s)
                s_lower = s.lower()
                return s == '' or s_lower in ['nan', 'none', 'null', '없음', '0', 'na', 'n/a', '미입력']
            결측치행 = df[df[실제필수컬럼].applymap(is_missing).any(axis=1)]
            if not 결측치행.empty:
                out_dir = os.path.dirname(input_file)
                out_path = self.get_incremented_filename("결측치_오류_행", "xlsx", out_dir)
                self.save_with_yellow_header(결측치행, out_path)
                import tkinter.messagebox as mb
                mb.showwarning('결측치 경고', f'필수 정보({', '.join(필수키)}) 결측치가 있는 {len(결측치행)}개 행이 발견되어 결과에서 제외되고,\n"{out_path}" 파일로 저장되었습니다.')
        # 결측치 없는 행만 사용
        df = df[~df.index.isin(결측치행.index)]
        if '상세주소' in df.columns and address_column in df.columns:
            def make_full_addr(row):
                base = str(row[address_column]) if pd.notna(row[address_column]) else ''
                detail = str(row['상세주소']) if pd.notna(row['상세주소']) and str(row['상세주소']).strip() else ''
                base = base.rstrip(', ').strip()
                if detail:
                    return f"{base} {detail}".strip()
                else:
                    return base
            df['최종주소'] = df.apply(make_full_addr, axis=1)
            address_for_classification = '최종주소'
        else:
            address_for_classification = address_column
        same_day_rows, dawn_rows, regular_rows = [], [], []
        same_day_set = {str(int(str(z).strip())).zfill(5) for z in self.same_day_zipcodes}
        dawn_set = {str(int(str(z).strip())).zfill(5) for z in self.dawn_zipcodes}
        total_rows = len(df)
        for idx, row in df.iterrows():
            address = row[address_for_classification]
            if '위례' in str(address):
                dawn_rows.append(row)
                continue
            zipcode = self.get_zipcode_from_address(address)
            zc_str = str(int(zipcode)).zfill(5) if zipcode else ''
            if zc_str in same_day_set:
                same_day_rows.append(row)
            elif zc_str in dawn_set:
                dawn_rows.append(row)
            else:
                regular_rows.append(row)
            if progress_callback:
                progress = ((idx + 1) / total_rows) * 100
                progress_callback(progress)
        output_dir = os.path.dirname(input_file)
        self.last_output_path = None
        last_file = None
        # 당일배송
        if same_day_rows:
            df_same = pd.DataFrame(same_day_rows)
            if df_same.empty:
                import tkinter.messagebox as mb
                mb.showinfo("안내", "당일배송 가능지역이 없습니다.")
            else:
                # 수량 반복 처리
                expanded_rows = []
                for _, row in df_same.iterrows():
                    qty = 1
                    qty_val = find_column(row, '수량')
                    if qty_val:
                        try:
                            qty = int(qty_val)
                        except:
                            qty = 1
                    for _ in range(qty):
                        new_row = row.copy()
                        # 수량 컬럼 찾아서 1로 설정
                        for col in df_same.columns:
                            if '수량' in col:
                                new_row[col] = '1'
                        expanded_rows.append(new_row)
                df_same = pd.DataFrame(expanded_rows)
                
                if '최종주소' in df_same.columns:
                    df_same['받는분주소'] = df_same['최종주소'].apply(clean_address_for_output)
                    for col in COLUMN_MAP['수령인상세주소']:
                        if col in df_same.columns:
                            df_same[col] = ''
                    df_same = df_same.drop(columns=['최종주소'])
                elif '받는분주소' in df_same.columns:
                    df_same['받는분주소'] = df_same['받는분주소'].apply(clean_address_for_output)
                # 상세주소 컬럼 자체 삭제
                for col in COLUMN_MAP['수령인상세주소']:
                    if col in df_same.columns:
                        df_same = df_same.drop(columns=[col])
                # 전화번호 하이픈 적용 (모든 후보 컬럼)
                for col in COLUMN_MAP['수령인연락처1']:
                    if col in df_same.columns:
                        df_same[col] = df_same[col].apply(format_phone_number)
                for col in COLUMN_MAP['수령인연락처2']:
                    if col in df_same.columns:
                        df_same[col] = df_same[col].apply(format_phone_number)
                for col in COLUMN_MAP['구매자연락처']:
                    if col in df_same.columns:
                        df_same[col] = df_same[col].apply(format_phone_number)
                # '배송지점' 등 불필요 컬럼 일괄 삭제
                df_same = merge_address_columns(df_same)
                out_path = self.get_incremented_filename("당일배송_주문리스트", "xlsx", output_dir)
                self.save_with_yellow_header(df_same, out_path)
                self.last_output_path = out_path
                last_file = out_path
        # 새벽배송
        if dawn_rows:
            result = []
            template = pd.read_excel(os.path.join('data', '새벽배송양식.xlsx'))
            columns = list(template.columns)
            tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            for row in dawn_rows:
                qty = 1
                qty_val = find_column(row, '수량')
                if qty_val:
                    try:
                        qty = int(qty_val)
                    except:
                        qty = 1
                for _ in range(qty):
                    new_row = {col: '' for col in columns}
                    new_row['배송요청일*'] = tomorrow
                    new_row['거래처 주문번호'] = find_column(row, '거래처주문코드')
                    new_row['주문자*'] = find_column(row, '주문자')
                    new_row['수령자*'] = find_column(row, '수령자')
                    new_row['상품명*'] = self.make_dawn_product_name(row, chasu)
                    road_addr = clean_address_for_output(replace_company_name(find_column(row, '수령자도로명주소')))
                    detail_addr = find_column(row, '수령자상세주소')
                    if detail_addr and str(detail_addr).strip():
                        full_addr = f"{road_addr} {str(detail_addr).strip()}"
                    else:
                        full_addr = road_addr
                    new_row['수령자 도로명 주소*'] = full_addr.strip()
                    new_row['수령자 상세주소'] = ''
                    new_row['수령자 연락처*'] = format_phone_number(find_column(row, '수령자연락처'))
                    dawn_msg = find_column(row, '배송메시지')
                    def is_etc_condition(val):
                        if val is None:
                            return True
                        sval = str(val).strip()
                        if sval == '' or sval.lower() == 'nan' or sval == '0 / 0':
                            return True
                        return False
                    for col in columns:
                        if '배송 받을 장소' in col:
                            new_row[col] = '기타' if is_etc_condition(dawn_msg) else '문 앞'
                        if '배송 받을 장소 상세' in col:
                            new_row[col] = '문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송'
                        if '배송 유형' in col:
                            new_row[col] = dawn_type or '배송대행'
                        if '배송 문자 전송 시점' in col:
                            new_row[col] = sms_type or '즉시전송'
                        if '출입방법' in col.replace(' ', ''):
                            new_row[col] = dawn_msg
                    result.append(new_row)
            df_result = pd.DataFrame(result, columns=columns)
            if df_result.empty:
                import tkinter.messagebox as mb
                mb.showinfo("안내", "새벽배송 가능지역이 없습니다.")
                return None
            # 새벽배송에서는 상세주소 컬럼을 삭제하지 않음!
            df_result = merge_address_columns(df_result)
            out_path = self.get_incremented_filename("새벽배송_주문리스트", "xlsx", output_dir)
            self.save_with_yellow_header(df_result, out_path)
            self.last_output_path = out_path
            last_file = out_path
        # 택배배송
        if regular_rows:
            df_reg = pd.DataFrame(regular_rows)
            if df_reg.empty:
                import tkinter.messagebox as mb
                mb.showinfo("안내", "택배배송 가능지역이 없습니다.")
                return None
            # 정상 주소만 필터링 (우편번호를 찾을 수 있는 경우만)
            valid_rows = []
            for _, row in df_reg.iterrows():
                address = row[address_for_classification]
                zipcode = self.get_zipcode_from_address(address)
                if zipcode:  # 우편번호를 찾을 수 있는 경우만 택배로 처리
                    valid_rows.append(row)
            
            if valid_rows:
                df_reg = pd.DataFrame(valid_rows)
                # 수량 반복 처리
                expanded_rows = []
                for _, row in df_reg.iterrows():
                    qty = 1
                    qty_val = find_column(row, '수량')
                    if qty_val:
                        try:
                            qty = int(qty_val)
                        except:
                            qty = 1
                    for _ in range(qty):
                        new_row = row.copy()
                        # 수량 컬럼 찾아서 1로 설정
                        for col in df_reg.columns:
                            if '수량' in col:
                                new_row[col] = '1'
                        expanded_rows.append(new_row)
                df_reg = pd.DataFrame(expanded_rows)
                
                if '최종주소' in df_reg.columns:
                    df_reg['받는분주소'] = df_reg['최종주소'].apply(clean_address_for_output)
                    for col in COLUMN_MAP['수령인상세주소']:
                        if col in df_reg.columns:
                            df_reg[col] = ''
                    df_reg = df_reg.drop(columns=['최종주소'])
                elif '받는분주소' in df_reg.columns:
                    df_reg['받는분주소'] = df_reg['받는분주소'].apply(clean_address_for_output)
                # 상세주소 컬럼 자체 삭제
                for col in COLUMN_MAP['수령인상세주소']:
                    if col in df_reg.columns:
                        df_reg = df_reg.drop(columns=[col])
                # 전화번호 하이픈 적용 (모든 후보 컬럼)
                for col in COLUMN_MAP['수령인연락처1']:
                    if col in df_reg.columns:
                        df_reg[col] = df_reg[col].apply(format_phone_number)
                for col in COLUMN_MAP['수령인연락처2']:
                    if col in df_reg.columns:
                        df_reg[col] = df_reg[col].apply(format_phone_number)
                for col in COLUMN_MAP['구매자연락처']:
                    if col in df_reg.columns:
                        df_reg[col] = df_reg[col].apply(format_phone_number)
                # '배송지점' 등 불필요 컬럼 일괄 삭제
                df_reg = merge_address_columns(df_reg)
                out_path = self.get_incremented_filename("택배배송_주문리스트", "xlsx", output_dir)
                self.save_with_yellow_header(df_reg, out_path)
                self.last_output_path = out_path
                last_file = out_path
        # 미분류 데이터 저장 (전체분류에서는 주소오류만 저장)
        if regular_rows:
            df_reg = pd.DataFrame(regular_rows)
            address_error_rows = []
            for _, row in df_reg.iterrows():
                address = row[address_for_classification]
                zipcode = self.get_zipcode_from_address(address)
                if not zipcode:  # 우편번호를 찾을 수 없는 경우만 주소오류로 처리
                    address_error_rows.append(row)
            if address_error_rows:
                df_addrerr = pd.DataFrame(address_error_rows)
                if df_addrerr.empty:
                    import tkinter.messagebox as mb
                    mb.showinfo("안내", "주소오류 데이터가 없습니다.")
                else:
                    # 상세주소 컬럼 제거
                    for col in COLUMN_MAP['수령인상세주소']:
                        if col in df_addrerr.columns:
                            df_addrerr = df_addrerr.drop(columns=[col])
                    # 최종주소가 있으면 받는분주소로 설정
                    if '최종주소' in df_addrerr.columns:
                        df_addrerr['받는분주소'] = df_addrerr['최종주소']
                        df_addrerr = df_addrerr.drop(columns=['최종주소'])
                    # 임시 컬럼 삭제
                    for col in ['_clean_name', '_clean_addr']:
                        if col in df_addrerr.columns:
                            df_addrerr = df_addrerr.drop(columns=[col])
                    out_path = self.get_incremented_filename("주소오류_미분류", "xlsx", output_dir)
                    self.save_with_yellow_header(df_addrerr, out_path)
                    self.last_output_path = out_path
                    last_file = out_path
        return last_file

    def classify_only(self, input_file: str, address_column: str, target: str, dawn_type=None, sms_type=None, chasu=None, progress_callback=None):
        df = read_excel_file(input_file, dtype=str)
        # 필수 컬럼 결측치 처리
        필수키 = ['구매자명', '구매자연락처1', '상품주문번호', '송하인주소', '수취인', '수취인연락처1', '받는분주소', '상품명', '수량', '수령인', '수령자']
        def normalize_colname(s):
            import re
            return re.sub(r'[\\s\\r\\n\\t\\u200b\\u3000]', '', str(s)).lower()
        실제필수컬럼 = set()
        norm_excel_cols = [normalize_colname(c) for c in df.columns]
        for key in 필수키:
            candidates = COLUMN_MAP.get(key, [key])
            for cand in candidates:
                norm_cand = normalize_colname(cand)
                for col, norm_col in zip(df.columns, norm_excel_cols):
                    if norm_col == norm_cand:
                        실제필수컬럼.add(col)
        실제필수컬럼 = list(실제필수컬럼)
        print('[디버그] 실제필수컬럼:', 실제필수컬럼)
        print('[디버그] df.columns:', list(df.columns))
        for col in 실제필수컬럼:
            try:
                print(f'[디버그] [{col}] 값들:', df[col].values)
            except Exception as e:
                print(f'[디버그] [{col}] 값 출력 오류:', e)
        if 실제필수컬럼:
            def is_missing(val):
                import pandas as pd, re
                if pd.isnull(val):
                    return True
                s = str(val).strip()
                s = re.sub(r'[\s\r\n\t\u200b\u3000]', '', s)
                s_lower = s.lower()
                return s == '' or s_lower in ['nan', 'none', 'null', '없음', '0', 'na', 'n/a', '미입력']
            결측치행 = df[df[실제필수컬럼].applymap(is_missing).any(axis=1)]
            if not 결측치행.empty:
                out_dir = os.path.dirname(input_file)
                out_path = self.get_incremented_filename("결측치_오류_행", "xlsx", out_dir)
                self.save_with_yellow_header(결측치행, out_path)
                import tkinter.messagebox as mb
                mb.showwarning('결측치 경고', f'필수 정보({', '.join(필수키)}) 결측치가 있는 {len(결측치행)}개 행이 발견되어 결과에서 제외되고,\n"{out_path}" 파일로 저장되었습니다.')
        # 결측치 없는 행만 사용
        df = df[~df.index.isin(결측치행.index)]
        if '상세주소' in df.columns and address_column in df.columns:
            def make_full_addr(row):
                base = str(row[address_column]) if pd.notna(row[address_column]) else ''
                detail = str(row['상세주소']) if pd.notna(row['상세주소']) and str(row['상세주소']).strip() else ''
                return base if not detail else base + ' ' + detail
            df['최종주소'] = df.apply(make_full_addr, axis=1)
            address_for_classification = '최종주소'
        else:
            address_for_classification = address_column
        rows = []
        unclassified_rows = []
        address_error_rows = []
        same_day_set = {str(int(str(z).strip())).zfill(5) for z in self.same_day_zipcodes}
        dawn_set = {str(int(str(z).strip())).zfill(5) for z in self.dawn_zipcodes}
        total_rows = len(df)
        for idx, row in df.iterrows():
            if progress_callback:
                progress = int((idx + 1) / total_rows * 100)
                progress_callback(progress)
            address = row[address_for_classification]
            if target == 'dawn' and '위례' in str(address):
                rows.append(row)
                continue
            zipcode = self.get_zipcode_from_address(address)
            zc_str = str(int(zipcode)).zfill(5) if zipcode else ''
            if not zipcode:
                if target == 'dawn':
                    address_error_rows.append(row)
                continue
            if target == 'same_day' and zc_str in same_day_set:
                rows.append(row)
            elif target == 'dawn' and zc_str in dawn_set:
                rows.append(row)
            elif target == 'dawn':
                unclassified_rows.append(row)
        output_dir = os.path.dirname(input_file)
        last_file = None
        if target == 'same_day':
            df_result = pd.DataFrame(rows)
            if df_result.empty:
                import tkinter.messagebox as mb
                mb.showinfo("안내", "당일배송 가능지역이 없습니다.")
                return None
            # 수량 반복 처리
            expanded_rows = []
            for _, row in df_result.iterrows():
                qty = 1
                qty_val = find_column(row, '수량')
                if qty_val:
                    try:
                        qty = int(qty_val)
                    except:
                        qty = 1
                for _ in range(qty):
                    new_row = row.copy()
                    # 수량 컬럼 찾아서 1로 설정
                    for col in df_result.columns:
                        if '수량' in col:
                            new_row[col] = '1'
                    expanded_rows.append(new_row)
            df_result = pd.DataFrame(expanded_rows)
            
            if '최종주소' in df_result.columns:
                df_result['받는분주소'] = df_result['최종주소'].apply(clean_address_for_output)
                for col in COLUMN_MAP['수령인상세주소']:
                    if col in df_result.columns:
                        df_result[col] = ''
                df_result = df_result.drop(columns=['최종주소'])
            # 상세주소 컬럼 자체 삭제 (새벽배송만 분류가 아닐 때만)
            for col in COLUMN_MAP['수령인상세주소']:
                if col in df_result.columns:
                    df_result = df_result.drop(columns=[col])
            out_path = self.get_incremented_filename("당일배송_주문리스트(단일)", "xlsx", output_dir)
            self.save_with_yellow_header(df_result, out_path)
            last_file = out_path
            # 미분류/주소오류 데이터 저장
            unclassified_rows = []
            address_error_rows = []
            for _, row in df.iterrows():
                address = row[address_for_classification]
                zipcode = self.get_zipcode_from_address(address)
                zc_str = str(int(zipcode)).zfill(5) if zipcode else ''
                if not zipcode:
                    address_error_rows.append(row)
                elif zc_str not in same_day_set:
                    unclassified_rows.append(row)
            if unclassified_rows:
                df_unclassified = pd.DataFrame(unclassified_rows)
                if df_unclassified.empty:
                    import tkinter.messagebox as mb
                    mb.showinfo("안내", "미분류 데이터가 없습니다.")
                else:
                    df_unclassified = merge_address_columns(df_unclassified)
                    out_path = self.get_incremented_filename("미분류_주문리스트(당일불가지역)", "xlsx", output_dir)
                    self.save_with_yellow_header(df_unclassified, out_path)
            if address_error_rows:
                df_addrerr = pd.DataFrame(address_error_rows)
                df_addrerr = merge_address_columns(df_addrerr)
                out_path = self.get_incremented_filename("주소오류_미분류", "xlsx", output_dir)
                self.save_with_yellow_header(df_addrerr, out_path)
        elif target == 'dawn':
            template = pd.read_excel(os.path.join('data', '새벽배송양식.xlsx'))
            columns = list(template.columns)
            tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            dawn_type = dawn_type or '배송대행'
            sms_type = sms_type or '즉시전송'
            result = []
            for row in rows:
                qty = 1
                if '수량' in row and pd.notna(row['수량']):
                    try:
                        qty = int(row['수량'])
                    except:
                        qty = 1
                for _ in range(qty):
                    new_row = {col: '' for col in columns}
                    new_row['배송요청일*'] = tomorrow
                    new_row['거래처 주문번호'] = find_column(row, '거래처주문코드')
                    new_row['주문자*'] = find_column(row, '주문자')
                    new_row['수령자*'] = find_column(row, '수령자')
                    new_row['상품명*'] = self.make_dawn_product_name(row, chasu)
                    road_addr = clean_address_for_output(replace_company_name(find_column(row, '수령자도로명주소')))
                    detail_addr = find_column(row, '수령자상세주소')
                    if detail_addr and str(detail_addr).strip():
                        full_addr = f"{road_addr} {str(detail_addr).strip()}"
                    else:
                        full_addr = road_addr
                    new_row['수령자 도로명 주소*'] = full_addr.strip()
                    new_row['수령자 상세주소'] = ''
                    new_row['수령자 연락처*'] = format_phone_number(find_column(row, '수령자연락처'))
                    dawn_msg = find_column(row, '배송메시지')
                    def is_etc_condition(val):
                        if val is None:
                            return True
                        sval = str(val).strip()
                        if sval == '' or sval.lower() == 'nan' or sval == '0 / 0':
                            return True
                        return False
                    for col in columns:
                        if '배송 받을 장소' in col:
                            new_row[col] = '기타' if is_etc_condition(dawn_msg) else '문 앞'
                        if '배송 받을 장소 상세' in col:
                            new_row[col] = '문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송'
                        if '배송 유형' in col:
                            new_row[col] = dawn_type or '배송대행'
                        if '배송 문자 전송 시점' in col:
                            new_row[col] = sms_type or '즉시전송'
                        if '출입방법' in col.replace(' ', ''):
                            new_row[col] = dawn_msg
                    result.append(new_row)
            if not result:
                import tkinter.messagebox as mb
                mb.showinfo("안내", "새벽배송 가능지역이 없습니다.")
                return None
            df_result = pd.DataFrame(result, columns=columns)
            if '최종주소' in df_result.columns:
                df_result['받는분주소'] = df_result['최종주소'].apply(clean_address_for_output)
                for col in COLUMN_MAP['수령인상세주소']:
                    if col in df_result.columns:
                        df_result[col] = ''
                df_result = df_result.drop(columns=['최종주소'])
            # 새벽배송만 분류에서는 상세주소 컬럼을 삭제하지 않음
            out_path = self.get_incremented_filename("새벽배송_주문리스트(단일)", "xlsx", output_dir)
            self.save_with_yellow_header(df_result, out_path, header=columns)
            last_file = out_path
            # 미분류 데이터 저장
            if unclassified_rows:
                df_unclassified = pd.DataFrame(unclassified_rows)
                if df_unclassified.empty:
                    import tkinter.messagebox as mb
                    mb.showinfo("안내", "미분류 데이터가 없습니다.")
                else:
                    df_unclassified = merge_address_columns(df_unclassified)
                    out_path = self.get_incremented_filename("미분류_주문리스트", "xlsx", output_dir)
                    self.save_with_yellow_header(df_unclassified, out_path)
            # 주소오류 데이터 저장
            if address_error_rows:
                df_addrerr = pd.DataFrame(address_error_rows)
                df_addrerr = merge_address_columns(df_addrerr)
                out_path = self.get_incremented_filename("주소오류_미분류", "xlsx", output_dir)
                self.save_with_yellow_header(df_addrerr, out_path)
        return last_file

    def classify_day_and_dawn(self, input_file: str, address_column: str, progress_callback=None):
        df = read_excel_file(input_file, dtype=str)
        # 필수 컬럼 결측치 처리
        필수키 = ['구매자명', '구매자연락처1', '상품주문번호', '송하인주소', '수취인', '수취인연락처1', '받는분주소', '상품명', '수량', '수령인', '수령자']
        def normalize_colname(s):
            import re
            return re.sub(r'[\\s\\r\\n\\t\\u200b\\u3000]', '', str(s)).lower()
        실제필수컬럼 = set()
        norm_excel_cols = [normalize_colname(c) for c in df.columns]
        for key in 필수키:
            candidates = COLUMN_MAP.get(key, [key])
            for cand in candidates:
                norm_cand = normalize_colname(cand)
                for col, norm_col in zip(df.columns, norm_excel_cols):
                    if norm_col == norm_cand:
                        실제필수컬럼.add(col)
        실제필수컬럼 = list(실제필수컬럼)
        print('[디버그] 실제필수컬럼:', 실제필수컬럼)
        print('[디버그] df.columns:', list(df.columns))
        for col in 실제필수컬럼:
            try:
                print(f'[디버그] [{col}] 값들:', df[col].values)
            except Exception as e:
                print(f'[디버그] [{col}] 값 출력 오류:', e)
        if 실제필수컬럼:
            def is_missing(val):
                import pandas as pd, re
                if pd.isnull(val):
                    return True
                s = str(val).strip()
                s = re.sub(r'[\s\r\n\t\u200b\u3000]', '', s)
                s_lower = s.lower()
                return s == '' or s_lower in ['nan', 'none', 'null', '없음', '0', 'na', 'n/a', '미입력']
            결측치행 = df[df[실제필수컬럼].applymap(is_missing).any(axis=1)]
            if not 결측치행.empty:
                out_dir = os.path.dirname(input_file)
                out_path = self.get_incremented_filename("결측치_오류_행", "xlsx", out_dir)
                self.save_with_yellow_header(결측치행, out_path)
                import tkinter.messagebox as mb
                mb.showwarning('결측치 경고', f'필수 정보({', '.join(필수키)}) 결측치가 있는 {len(결측치행)}개 행이 발견되어 결과에서 제외되고,\n"{out_path}" 파일로 저장되었습니다.')
        # 결측치 없는 행만 사용
        df = df[~df.index.isin(결측치행.index)]
        if '상세주소' in df.columns and address_column in df.columns:
            def make_full_addr(row):
                base = str(row[address_column]) if pd.notna(row[address_column]) else ''
                detail = str(row['상세주소']) if pd.notna(row['상세주소']) and str(row['상세주소']).strip() else ''
                return base if not detail else base + ' ' + detail
            df['최종주소'] = df.apply(make_full_addr, axis=1)
            address_for_classification = '최종주소'
        else:
            address_for_classification = address_column
        same_day_rows, dawn_rows = [], []
        dawn_indices = []
        same_day_set = {str(int(str(z).strip())).zfill(5) for z in self.same_day_zipcodes}
        dawn_set = {str(int(str(z).strip())).zfill(5) for z in self.dawn_zipcodes}
        
        total_rows = len(df)
        for idx, row in df.iterrows():
            if progress_callback:
                progress = int((idx + 1) / total_rows * 100)
                progress_callback(progress)
                
            address = row[address_for_classification]
            if '위례' in str(address):
                dawn_rows.append(row)
                dawn_indices.append(idx)
                continue
            zipcode = self.get_zipcode_from_address(address)
            zc_str = str(int(zipcode)).zfill(5) if zipcode else ''
            if zc_str in same_day_set:
                same_day_rows.append(row)
            elif zc_str in dawn_set:
                dawn_rows.append(row)
                
        if progress_callback:
            progress = 100
            progress_callback(progress)
            
        combined_rows = same_day_rows + dawn_rows
        # 순서 보장: 원본 인덱스 순회하며 조건에 맞는 행만 추출
        filtered_rows = []
        address_error_rows = []
        for idx, row in df.iterrows():
            address = row[address_for_classification]
            zipcode = self.get_zipcode_from_address(address)
            zc_str = str(int(zipcode)).zfill(5) if zipcode else ''
            if not zipcode:
                address_error_rows.append(row)
                continue
            # 분류 우선순위: 당일 > 새벽
            if zc_str in same_day_set:
                filtered_rows.append((idx, row, ''))  # 당일: ''
            elif zc_str in dawn_set:
                filtered_rows.append((idx, row, '108'))  # 새벽: 108
        # dawn_indices에 포함된 인덱스는 filtered_rows에서 제외 (중복 방지)
        filtered_rows = [(idx, row, code) for (idx, row, code) in filtered_rows if idx not in dawn_indices]
        # 수량 반복 처리 및 집하지점코드 추가
        expanded_rows = []
        # 1. 기존 filtered_rows(우편번호 기반) 처리
        for idx, row, code in filtered_rows:
            qty = 1
            qty_val = find_column(row, '수량')
            if qty_val:
                try:
                    qty = int(qty_val)
                except:
                    qty = 1
            for _ in range(qty):
                new_row = row.copy()
                for col in df.columns:
                    if '수량' in col:
                        new_row[col] = '1'
                new_row['집하지점코드'] = code
                expanded_rows.append((idx, new_row))
        # 2. '위례' 포함된 주소도 반드시 집하지점코드 108로 추가
        for idx in dawn_indices:
            row = df.iloc[idx]
            qty = 1
            qty_val = find_column(row, '수량')
            if qty_val:
                try:
                    qty = int(qty_val)
                except:
                    qty = 1
            for _ in range(qty):
                new_row = row.copy()
                for col in df.columns:
                    if '수량' in col:
                        new_row[col] = '1'
                new_row['집하지점코드'] = '108'
                expanded_rows.append((idx, new_row))
        # 원본 순서대로 정렬
        expanded_rows.sort(key=lambda x: x[0])
        combined_df = pd.DataFrame([row for idx, row in expanded_rows])
        # 주소 컬럼 클린업
        if combined_df.empty:
            import tkinter.messagebox as mb
            mb.showinfo("안내", "당일배송/새벽배송 대상이 없습니다.")
            return None
        
        # 원본 데이터에서 주소 + 상세주소 합치기 처리
        def normalize_colname(s):
            return re.sub(r'[\s\-_]+', '', str(s))
        
        # 상세주소 컬럼 찾기
        detail_col = None
        for col in combined_df.columns:
            norm_col = normalize_colname(col)
            if any(keyword in norm_col for keyword in ['수령인상세주소', '받는분상세주소', '상세주소']):
                detail_col = col
                break
        
        # 주소 합치기 로직 (당일배송만 분류와 동일)
        if detail_col and detail_col in combined_df.columns:
            def make_full_addr(row):
                base = str(row[address_column]) if pd.notna(row[address_column]) else ''
                detail = str(row[detail_col]) if pd.notna(row[detail_col]) and str(row[detail_col]).strip() and str(row[detail_col]).strip().lower() != 'nan' else ''
                base = base.rstrip(', ').strip()
                if detail:
                    return f"{base} {detail}".strip()
                else:
                    return base
            combined_df['최종주소'] = combined_df.apply(make_full_addr, axis=1)
            address_for_classification = '최종주소'
        else:
            address_for_classification = address_column
        
        # 최종주소를 받는분주소에 복사하고 정리
        if '최종주소' in combined_df.columns:
            combined_df[address_column] = combined_df['최종주소'].apply(clean_address_for_output)
            for col in COLUMN_MAP['수령인상세주소']:
                if col in combined_df.columns:
                    combined_df[col] = ''
            combined_df = combined_df.drop(columns=['최종주소'])
        elif address_column in combined_df.columns:
            combined_df[address_column] = combined_df[address_column].apply(clean_address_for_output)
        
        # 상세주소 컬럼 자체 삭제
        for col in COLUMN_MAP['수령인상세주소']:
            if col in combined_df.columns:
                combined_df = combined_df.drop(columns=[col])
        # 전화번호 하이픈 적용 (모든 후보 컬럼)
        for col in COLUMN_MAP['수령인연락처1']:
            if col in combined_df.columns:
                combined_df[col] = combined_df[col].apply(format_phone_number)
        for col in COLUMN_MAP['수령인연락처2']:
            if col in combined_df.columns:
                combined_df[col] = combined_df[col].apply(format_phone_number)
        for col in COLUMN_MAP['구매자연락처']:
            if col in combined_df.columns:
                combined_df[col] = combined_df[col].apply(format_phone_number)
        # 결과에서 '배송지점' 컬럼이 있으면 삭제
        if '배송지점' in combined_df.columns:
            combined_df = combined_df.drop(columns=['배송지점'])
        output_dir = os.path.dirname(input_file)
        output_filename = "당일새벽배송_주문리스트.xlsx"
        out_path = self.get_incremented_filename(output_filename, "xlsx", output_dir)
        self.save_with_yellow_header(combined_df, out_path)
        # 주소오류 데이터 저장
        if address_error_rows:
            df_addrerr = pd.DataFrame(address_error_rows)
            df_addrerr = merge_address_columns(df_addrerr)  # 상세주소 합치기 및 정리
            out_path = self.get_incremented_filename("주소오류_미분류", "xlsx", output_dir)
            self.save_with_yellow_header(df_addrerr, out_path)
        return out_path

    # 상품명 생성 함수 (차수 반영)
    @staticmethod
    def make_dawn_product_name(row, chasu=None):
        # chasu가 '차수없음'이거나 None/빈값이면 상품명에 차수 미포함
        if chasu in [None, '', '차수없음']:
            chasu_str = ''
        else:
            chasu_str = f" {chasu}"
        buyer = find_column(row, '구매자명') or find_column(row, '주문자')
        product = find_column(row, '상품명')
        order_addr = find_column(row, '주문자주소')
        order_code = find_column(row, '거래처주문코드')
        
        # 주문자명 마스킹 처리 (첫글자와 마지막글자만 표시, 중간은 *로 대체)
        masked_buyer = ''
        if buyer and str(buyer).lower() != 'nan' and buyer is not None:
            buyer_str = str(buyer)
            if len(buyer_str) == 2:  # 두 글자 이름인 경우 (예: 홍길)
                masked_buyer = buyer_str[0] + '*'
            elif len(buyer_str) >= 3:  # 세 글자 이상 이름인 경우 (예: 홍길동, 홍길동순)
                masked_buyer = buyer_str[0] + '*' * (len(buyer_str) - 2) + buyer_str[-1]
            else:  # 한 글자 이름인 경우
                masked_buyer = buyer_str
        else:
            masked_buyer = ''
            
        prefix = "주문 : " + masked_buyer
        
        # 주소, 주문자, 거래처주문코드 등에서 남촌상회 포함 여부 체크
        def safe_str(val):
            return str(val) if val is not None else ''
        if order_addr:
            cleaned_addr = re.sub(r'[\s\-_]+', '', safe_str(order_addr))
            if '이낙근' in cleaned_addr:
                prefix += " 이낙근"
            if '국앤찬' in cleaned_addr:
                prefix += " 국앤찬"
            if has_jonggakne_company(order_addr):
                prefix += " 총각네"
            if '맛사랑' in cleaned_addr:
                prefix += " 맛사랑"
            if '새농' in cleaned_addr:
                prefix += " 새농"
            if '당일장터' in cleaned_addr:
                prefix += " 당일장터"
            if '남촌상회' in cleaned_addr or '남촌과일' in cleaned_addr:
                prefix += " 남촌상회"
        # 거래처주문코드(상품주문번호 등)에도 남촌상회가 있으면 추가
        if order_code and ('남촌상회' in safe_str(order_code) or '남촌과일' in safe_str(order_code)):
            prefix += " 남촌상회"
        # 구매자명/주문자명에도 남촌상회가 있으면 추가
        if buyer and ('남촌상회' in safe_str(buyer) or '남촌과일' in safe_str(buyer)):
            prefix += " 남촌상회"
        result = f"{prefix}{chasu_str} {product}".strip()
        return result

    def classify_all_delivery_types(self, df, output_dir=None, input_file_path=None, progress_callback=None):
        # 주소 컬럼 자동 인식
        address_col = None
        for col in ADDRESS_CANDIDATES:
            if col in df.columns:
                address_col = col
                break
        if not address_col:
            raise KeyError('주소 컬럼을 찾을 수 없습니다.')
        same_day_set = {str(int(str(z).strip())).zfill(5) for z in self.same_day_zipcodes}
        dawn_set = {str(int(str(z).strip())).zfill(5) for z in self.dawn_zipcodes}
        result_rows = []
        address_error_rows = []
        total_rows = len(df)
        for idx, row in df.iterrows():
            if progress_callback:
                progress = int((idx + 1) / total_rows * 100)
                progress_callback(progress)
            address = row[address_col]
            if '위례' in str(address):
                qty = 1
                qty_val = None
                for qcol in COLUMN_MAP['수량']:
                    if qcol in df.columns:
                        qty_val = row[qcol]
                        break
                if qty_val:
                    try:
                        qty = int(qty_val)
                    except:
                        qty = 1
                for _ in range(qty):
                    new_row = row.copy()
                    new_row['집하지점코드'] = '108'
                    for qcol in COLUMN_MAP['수량']:
                        if qcol in new_row:
                            new_row[qcol] = '1'
                    result_rows.append(new_row)
                continue
            zipcode = self.get_zipcode_from_address(address)
            zc_str = str(int(zipcode)).zfill(5) if zipcode else ''
            qty = 1
            qty_val = None
            for qcol in COLUMN_MAP['수량']:
                if qcol in df.columns:
                    qty_val = row[qcol]
                    break
            if qty_val:
                try:
                    qty = int(qty_val)
                except:
                    qty = 1
            if not zipcode:
                address_error_rows.append(row)
                continue
            for _ in range(qty):
                new_row = row.copy()
                if zc_str in same_day_set:
                    new_row['집하지점코드'] = ''  # 당일배송
                elif zc_str in dawn_set:
                    new_row['집하지점코드'] = '108'  # 새벽배송
                else:
                    new_row['집하지점코드'] = '105'  # 택배
                for qcol in COLUMN_MAP['수량']:
                    if qcol in new_row:
                        new_row[qcol] = '1'
                result_rows.append(new_row)
        result_df = pd.DataFrame(result_rows, columns=df.columns.tolist() + ['집하지점코드'] if '집하지점코드' not in df.columns else df.columns)
        result_df = merge_address_columns(result_df)
        # 결과에서 '배송지점' 컬럼이 있으면 삭제
        if '배송지점' in result_df.columns:
            result_df = result_df.drop(columns=['배송지점'])
        if address_error_rows and input_file_path:
            out_dir = output_dir or os.path.dirname(input_file_path)
            df_addrerr = pd.DataFrame(address_error_rows)
            df_addrerr = merge_address_columns(df_addrerr)
            out_path = self.get_incremented_filename("주소오류_미분류", "xlsx", out_dir)
            self.save_with_yellow_header(df_addrerr, out_path)
        if input_file_path:
            out_dir = output_dir or os.path.dirname(input_file_path)
            base = os.path.splitext(os.path.basename(input_file_path))[0] + '_당일새벽택배한방양식'
            out_path = self.get_incremented_filename(base, 'xlsx', out_dir)
            self.save_with_yellow_header(result_df, out_path)
            self.last_output_path = out_path
            return out_path
        return None

def merge_address_columns(df):
    return df

class DeliveryClassifierGUI:
    def __init__(self):
        self.root = tk.Tk()  # 반드시 가장 먼저!
        self.root.title("허우적 배송 분류 프로그램")
        self.root.geometry("900x800")  # 기존 700 → 900으로 확대
        # 로고 이미지 안전하게 불러오기 (data/logo.png)
        try:
            from PIL import Image, ImageTk
            logo_path = os.path.join('data', 'logo.png')
            if os.path.exists(logo_path):
                logo_img = Image.open(logo_path)
                logo_img.thumbnail((80, 80), Image.LANCZOS)  # 기존 50 → 80으로 확대
                self.logo = ImageTk.PhotoImage(logo_img)
            else:
                print(f"[경고] 로고 이미지 파일이 존재하지 않습니다: {logo_path}")
                self.logo = None
        except Exception as e:
            print(f"[경고] 로고 이미지 로딩 실패: {e}")
            self.logo = None
        self.classifier = DeliveryClassifier()
        self.input_file = ""
        self.address_column = ""
        self.progress_var = tk.DoubleVar()
        self.progress_bar = None
        self.is_processing = False
        self.admin_authenticated = False
        self.api_key_visible = False
        self.naver_green = '#03C75A'
        self.bg_color = '#f7f7f7'
        self.root.configure(bg=self.bg_color)
        self.set_style()
        self.create_tabs()
        self.check_api_key()

    def set_style(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TButton',
            font=('맑은 고딕', 10, 'bold'),
            background='#f8f8f8',
            foreground='#222',
            borderwidth=2,
            focusthickness=1,
            focuscolor='#03C75A',
            padding=8,
            relief='raised')
        style.map('TButton',
            background=[('active', '#eafff3'), ('!active', '#f8f8f8')],
            foreground=[('active', '#03C75A'), ('!active', '#222')],
            bordercolor=[('active', '#03C75A'), ('!active', '#b0b0b0')])
        style.configure('TLabel', font=('맑은 고딕', 10), background=self.bg_color)
        style.configure('TEntry', font=('맑은 고딕', 10), fieldbackground='white', bordercolor='#d0d0d0')
        style.configure('TFrame', background=self.bg_color)
        style.configure('TLabelframe', background=self.bg_color, font=('맑은 고딕', 10, 'bold'))
        style.configure('TLabelframe.Label', background=self.bg_color, font=('맑은 고딕', 10, 'bold'))
        style.configure('Horizontal.TProgressbar', troughcolor='#e0e0e0', background=self.naver_green, thickness=12)

    def create_tabs(self):
        self.tab_control = ttk.Notebook(self.root)
        self.tab_main = ttk.Frame(self.tab_control, style='TFrame')
        self.tab_chasu = ttk.Frame(self.tab_control, style='TFrame')  # 새벽차수추가 탭
        self.tab_invoice = ttk.Frame(self.tab_control, style='TFrame')  # 송장매칭 탭
        self.tab_zipcode = ttk.Frame(self.tab_control, style='TFrame')
        self.tab_apikey = ttk.Frame(self.tab_control, style='TFrame')
        self.tab_data = ttk.Frame(self.tab_control, style='TFrame')  # 데이터 관리 탭
        self.tab_control.add(self.tab_main, text='배송 분류')
        self.tab_control.add(self.tab_chasu, text='새벽옵션추가')
        self.tab_control.add(self.tab_invoice, text='송장매칭')
        self.tab_control.add(self.tab_zipcode, text='우편번호 관리')
        self.tab_control.add(self.tab_apikey, text='API 키 입력/변경')
        self.tab_control.add(self.tab_data, text='데이터 관리')
        self.tab_control.pack(expand=1, fill='both')
        self.create_main_tab()
        self.create_zipcode_tab()
        self.create_apikey_tab()
        self.create_chasu_tab()
        self.create_data_tab()
        self.create_invoice_tab()
        self.tab_control.bind('<<NotebookTabChanged>>', self.on_tab_changed)

    def on_tab_changed(self, event):
        selected_tab = event.widget.select()
        tab_text = event.widget.tab(selected_tab, "text")
        if tab_text == 'API 키 입력/변경' and not self.admin_authenticated:
            pw = tk.simpledialog.askstring("관리자 인증", "관리자 비밀번호를 입력하세요.", show="*")
            if pw == "admin1234":
                self.admin_authenticated = True
                self.create_apikey_tab()
            else:
                messagebox.showwarning("접근 제한", "관리자만 접근할 수 있습니다.")
                self.tab_control.select(self.tab_main)
        elif tab_text == '데이터 관리' and not self.admin_authenticated:
            pw = tk.simpledialog.askstring("관리자 인증", "관리자 비밀번호를 입력하세요.", show="*")
            if pw == "admin1234":
                self.admin_authenticated = True
                self.create_data_tab()
            else:
                messagebox.showwarning("접근 제한", "관리자만 접근할 수 있습니다.")
                self.tab_control.select(self.tab_main)

    def create_main_tab(self):
        frame = self.tab_main
        for widget in frame.winfo_children():
            widget.destroy()
        # 상단 로고 + 타이틀 (가운데 정렬)
        logo_title_frame = tk.Frame(frame, bg=self.bg_color)
        logo_title_frame.pack(pady=10)
        if self.logo:
            logo_label = tk.Label(logo_title_frame, image=self.logo, bg=self.bg_color)
            logo_label.image = self.logo  # 가비지 컬렉션 방지
            logo_label.pack(side=tk.LEFT, padx=(0, 18))  # 기존 10 → 18
        title_label = tk.Label(logo_title_frame, text="배송분류 자동화", font=('맑은 고딕', 20, 'bold'), bg=self.bg_color)
        title_label.pack(side=tk.LEFT)
        # 버전 표시
        version_label = tk.Label(frame, text=f"버전: {VERSION}", font=('맑은 고딕', 10), bg=self.bg_color, fg="#888")
        version_label.pack()
        subtitle_label = tk.Label(frame, text="주소기반 우편번호 매칭 배송방법 자동분류 프로그램", font=('맑은 고딕', 11), bg=self.bg_color)
        subtitle_label.pack()
        # 파일 선택
        file_frame = tk.Frame(frame, bg=self.bg_color)
        file_frame.pack(pady=10)
        self.file_label = tk.Label(file_frame, text="선택된 파일: 없음", font=('맑은 고딕', 10), bg=self.bg_color)
        self.file_label.pack(side=tk.LEFT, padx=5)
        self.btn_select = ttk.Button(file_frame, text="파일 선택", command=self.select_file)
        self.btn_select.pack(side=tk.LEFT, padx=5)
        # 안내문구
        tk.Label(frame, text="* 한방양식 업로드 시 주소 컬럼은 자동 인식 또는 직접 선택 가능합니다.", fg="#888", bg=self.bg_color, font=('맑은 고딕', 10)).pack(pady=1)
        # 새벽양식옵션 선택 그룹
        dawn_opt_frame = ttk.Labelframe(frame, text="새벽양식옵션 선택", padding=10, style='TLabelframe')
        dawn_opt_frame.pack(pady=8, fill="x", padx=35)
        tk.Label(
            dawn_opt_frame,
            text="*요청유형, 배송문자유형, 차수 를 변경 시 분류된 새벽양식에 모두 반영됩니다.\n미선택시 현재 세팅이 기본값입니다.",
            fg="#888", bg=self.bg_color, font=('맑은 고딕', 9),
            anchor="w", justify="left"
        ).pack(anchor="w", pady=(0, 6), fill="x")
        # 요청유형
        request_frame = tk.Frame(dawn_opt_frame, bg=self.bg_color)
        request_frame.pack(fill="x", pady=2)
        tk.Label(request_frame, text="요청유형:", font=('맑은 고딕', 10), bg=self.bg_color).pack(side=tk.LEFT, padx=5)
        self.request_type = ttk.Combobox(request_frame, values=["배송대행", "택배대행"], state="readonly", width=15)
        self.request_type.set("배송대행")
        self.request_type.pack(side=tk.LEFT, padx=5)
        # 배송문자유형
        sms_frame = tk.Frame(dawn_opt_frame, bg=self.bg_color)
        sms_frame.pack(fill="x", pady=2)
        tk.Label(sms_frame, text="배송문자유형:", font=('맑은 고딕', 10), bg=self.bg_color).pack(side=tk.LEFT, padx=5)
        self.sms_type = ttk.Combobox(sms_frame, values=["즉시전송", "7시전송"], state="readonly", width=15)
        self.sms_type.set("즉시전송")
        self.sms_type.pack(side=tk.LEFT, padx=5)
        # 차수 (10차까지)
        chasu_frame = tk.Frame(dawn_opt_frame, bg=self.bg_color)
        chasu_frame.pack(fill="x", pady=2)
        tk.Label(chasu_frame, text="차수:", font=('맑은 고딕', 10), bg=self.bg_color).pack(side=tk.LEFT, padx=5)
        self.chasu_var = tk.StringVar(value="차수없음")
        for chasu in ['차수없음'] + [f'{i}차' for i in range(1, 11)]:
            ttk.Radiobutton(chasu_frame, text=chasu, variable=self.chasu_var, value=chasu).pack(side=tk.LEFT, padx=2)
        # 배송 분류방법 선택 그룹
        classify_frame = ttk.Labelframe(frame, text="배송 분류방법 선택", padding=10, style='TLabelframe')
        classify_frame.pack(pady=8, fill="x", padx=35)
        tk.Label(classify_frame, text="1. 각 분류방법에 맞는 데이터로 자동추출됩니다.\n2. 주소오류 데이터 생성시 100%주소오류입니다. 정확한 도로명 주소를 확인해주세요\n3. 건물번호와 동호수 띄어쓰기 없이 붙어있을경우 빌게이츠랑 스티븐잡스가 와도 못찾습니다. 띄어쓰기 확인해주세요", fg="#888", bg=self.bg_color, font=('맑은 고딕', 9), justify="left").pack(anchor="w", pady=(0, 6))
        btn_row = tk.Frame(classify_frame, bg=self.bg_color)
        btn_row.pack()
        self.btn_all = ttk.Button(btn_row, text="전체 분류(당일→새벽→택배)", command=self.start_classification_all, width=26)
        self.btn_all.pack(side=tk.LEFT, padx=4)
        self.btn_day = ttk.Button(btn_row, text="당일배송만 분류", command=self.start_classification_same_day, width=20)
        self.btn_day.pack(side=tk.LEFT, padx=4)
        self.btn_dawn = ttk.Button(btn_row, text="새벽배송만 분류", command=self.start_classification_dawn, width=20)
        self.btn_dawn.pack(side=tk.LEFT, padx=4)

        # 일괄 구분방법 선택 그룹
        allinone_frame = ttk.Labelframe(frame, text="일괄 구분방법 선택", padding=10, style='TLabelframe')
        allinone_frame.pack(pady=8, fill="x", padx=35)
        tk.Label(allinone_frame, text="*일괄구분시 원본데이터 기준으로 집하지점코드 105,108 등이 자동구분생성됩니다.", fg="#888", bg=self.bg_color, font=('맑은 고딕', 9)).pack(anchor="w", pady=(0, 6))
        btn_row2 = tk.Frame(allinone_frame, bg=self.bg_color)
        btn_row2.pack()
        self.btn_day_and_dawn = ttk.Button(btn_row2, text="당일&새벽 한방양식 일괄구분", command=self.start_classification_day_and_dawn, width=26)
        self.btn_day_and_dawn.pack(side=tk.LEFT, padx=4)
        self.btn_all_in_one = ttk.Button(btn_row2, text="당일&새벽&택배 한방양식 일괄구분", command=self.start_classification_all_in_one, width=28)
        self.btn_all_in_one.pack(side=tk.LEFT, padx=4)

        # 진행상황 메시지
        self.progress_label = tk.Label(frame, text="", fg="#03C75A", bg=self.bg_color, font=('맑은 고딕', 11, 'bold'))
        self.progress_label.pack(pady=(5, 0))

    def create_zipcode_tab(self):
        frame = self.tab_zipcode
        # 당일배송
        tk.Label(frame, text="당일배송 우편번호 ({}개)".format(len(self.classifier.same_day_zipcodes))).pack(pady=5)
        tk.Button(frame, text="당일배송 우편번호 파일 교체", command=lambda: self.update_zipcode('same_day')).pack(pady=5)
        # 새벽배송
        tk.Label(frame, text="새벽배송 우편번호 ({}개)".format(len(self.classifier.dawn_zipcodes))).pack(pady=5)
        tk.Button(frame, text="새벽배송 우편번호 파일 교체", command=lambda: self.update_zipcode('dawn')).pack(pady=5)
        # 안내
        tk.Label(frame, text="* csv 파일은 한 줄에 하나의 우편번호만 있어야 합니다.").pack(pady=10)

    def create_apikey_tab(self):
        frame = self.tab_apikey
        for widget in frame.winfo_children():
            widget.destroy()
        outer = tk.Frame(frame)
        outer.pack(expand=True)
        tk.Label(outer, text="API 키 변경 (관리자 전용)", font=("맑은 고딕", 13, "bold")).pack(pady=(20, 10))
        # 현재 API 키 표시 (별표 처리, 눈동자 버튼)
        key_frame = tk.Frame(outer)
        key_frame.pack(pady=5)
        self.api_key_value = self.classifier.api_key or ''
        self.api_key_var = tk.StringVar()
        self.api_key_var.set('*' * len(self.api_key_value) if self.api_key_value else '')
        self.api_key_label = tk.Label(key_frame, textvariable=self.api_key_var, font=("Consolas", 12))
        self.api_key_label.pack(side=tk.LEFT)
        self.eye_btn = tk.Button(key_frame, text="👁️", command=self.toggle_api_key_visible, relief=tk.FLAT)
        self.eye_btn.pack(side=tk.LEFT, padx=5)
        tk.Label(outer, text="(현재 등록된 API 키)", fg="gray").pack()
        # 새 API 키 입력
        tk.Label(outer, text="새 API 키 입력:").pack(pady=(15, 2))
        self.api_key_entry = tk.Entry(outer, show="*", width=40)
        self.api_key_entry.pack(pady=2)
        # 변경 버튼
        tk.Button(outer, text="API 키 변경", command=self.confirm_change_api_key, width=20).pack(pady=10)
        self.api_key_status = tk.Label(outer, text="", fg="green")
        self.api_key_status.pack(pady=5)

    def toggle_api_key_visible(self):
        self.api_key_visible = not self.api_key_visible
        if self.api_key_visible:
            self.api_key_var.set(self.api_key_value)
            self.eye_btn.config(text="🙈")
        else:
            self.api_key_var.set('*' * len(self.api_key_value) if self.api_key_value else '')
            self.eye_btn.config(text="👁️")

    def confirm_change_api_key(self):
        new_key = self.api_key_entry.get().strip()
        if not new_key:
            self.api_key_status.config(text="새 API 키를 입력하세요.", fg="red")
            return
        if messagebox.askyesno("API 키 변경", "정말로 API 키를 변경하시겠습니까?"):
            self.save_api_key(new_key)

    def save_api_key(self, new_key=None):
        if new_key is None:
            new_key = self.api_key_entry.get().strip()
        if not new_key:
            self.api_key_status.config(text="API 키를 입력하세요.", fg="red")
            return
        self.classifier.save_api_key(new_key)
        self.api_key_value = new_key
        self.api_key_visible = False
        self.api_key_var.set('*' * len(new_key))
        self.eye_btn.config(text="👁️")
        self.api_key_status.config(text="API 키가 성공적으로 저장되었습니다.", fg="green")
        self.api_key_entry.delete(0, tk.END)

    def check_api_key(self):
        if not self.classifier.api_key:
            status = "API 키가 등록되어 있지 않습니다."
        else:
            status = "API 키가 정상적으로 등록되어 있습니다."
        if hasattr(self, 'api_key_status'):
            self.api_key_status.config(text=status, fg="green")

    def select_file(self):
        self.input_file = filedialog.askopenfilename(
            title="배송리스트 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if self.input_file:
            try:
                # Use the improved read_excel_file function
                df = read_excel_file(self.input_file)
                
                found_col = None
                for col in ADDRESS_CANDIDATES:
                    if col in df.columns:
                        found_col = col
                        break
                if found_col:
                    self.address_column = found_col
                    fname = os.path.basename(self.input_file)
                    self.file_label.config(text=f"선택: {fname} (주소컬럼: {found_col})")
                else:
                    import tkinter.simpledialog
                    col = tkinter.simpledialog.askstring(
                        "주소 컬럼명 입력",
                        f"주소로 사용할 컬럼명을 입력하세요.\n\n엑셀 컬럼명 목록:\n{', '.join(df.columns)}"
                    )
                    if col and col in df.columns:
                        self.address_column = col
                        fname = os.path.basename(self.input_file)
                        self.file_label.config(text=f"선택: {fname} (주소컬럼: {col})")
                    else:
                        messagebox.showerror("오류", "주소 컬럼을 찾을 수 없습니다. 파일을 확인해주세요.")
                        self.input_file = ""
                        self.file_label.config(text="(선택된 파일 없음)")
                        self.address_column = ""
            except Exception as e:
                messagebox.showerror("파일 읽기 오류", f"파일을 읽는 중 오류가 발생했습니다:\n{str(e)}")
                self.input_file = ""
                self.file_label.config(text="(선택된 파일 없음)")
                self.address_column = ""
        else:
            self.file_label.config(text="(선택된 파일 없음)")
            self.address_column = ""

    def show_progress(self):
        if not self.progress_bar:
            self.progress_label.config(text="분류 중입니다... 0%", fg="#03C75A")
            self.root.update_idletasks()
            self.progress_bar = ttk.Progressbar(self.tab_main, 
                                              variable=self.progress_var,
                                              maximum=100,
                                              mode='determinate',
                                              style="Custom.Horizontal.TProgressbar",
                                              length=300)
            self.progress_bar.pack(pady=15)
        self.progress_var.set(0)
        self.is_processing = True
        for btn in [getattr(self, n, None) for n in ['btn_all', 'btn_day_and_dawn', 'btn_day']]:
            if btn:
                btn.state(['disabled'])

    def update_progress(self, value):
        if self.progress_bar:
            self.progress_var.set(value)
            self.progress_label.config(text=f"분류 중입니다... {int(value)}%")
            self.root.update_idletasks()

    def threaded_classification(self, func, *args):
        import threading
        self.show_progress()
        def run():
            try:
                def progress_callback(progress):
                    self.root.after(0, lambda: self.update_progress(progress))
                import inspect
                params = list(inspect.signature(func).parameters)
                result = None
                if 'progress_callback' in params:
                    result = func(*args, progress_callback=progress_callback)
                else:
                    result = func(*args)
                # 분류 결과 파일 경로 추출
                file_path = result or getattr(self.classifier, 'last_output_path', None)
                msg = "분류가 완료되었습니다!"
                if file_path:
                    msg += f" | 저장 파일: {file_path}"
                self.root.after(0, lambda: self.show_custom_message("완료", msg, file_path))
            finally:
                self.root.after(0, self.hide_progress)
        threading.Thread(target=run, daemon=True).start()

    def hide_progress(self):
        if self.progress_bar:
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            self.progress_bar = None
        self.is_processing = False
        self.progress_label.config(text="")
        for btn in [getattr(self, n, None) for n in ['btn_all', 'btn_day_and_dawn', 'btn_day']]:
            if btn:
                btn.state(['!disabled'])

    def start_classification_all(self):
        if not self.input_file or not self.address_column:
            messagebox.showwarning("입력 오류", "파일과 주소 컬럼을 선택하세요.")
            return
        if self.is_processing:
            return
        chasu = self.chasu_var.get() if hasattr(self, 'chasu_var') else None
        dawn_type = self.request_type.get() if hasattr(self, 'request_type') else '배송대행'
        sms_type = self.sms_type.get() if hasattr(self, 'sms_type') else '즉시전송'
        self.threaded_classification(self.classifier.classify_all, self.input_file, self.address_column, chasu, dawn_type, sms_type)

    def start_classification_same_day(self):
        if not self.input_file or not self.address_column:
            messagebox.showwarning("입력 오류", "파일과 주소 컬럼을 선택하세요.")
            return
        if self.is_processing:
            return
        self.threaded_classification(self.classifier.classify_only, self.input_file, self.address_column, 'same_day')

    def start_classification_dawn(self):
        if not self.input_file or not self.address_column:
            messagebox.showwarning("입력 오류", "파일과 주소 컬럼을 선택하세요.")
            return
        if self.is_processing:
            return
        chasu = self.chasu_var.get() if hasattr(self, 'chasu_var') else None
        dawn_type = self.request_type.get() if hasattr(self, 'request_type') else '배송대행'
        sms_type = self.sms_type.get() if hasattr(self, 'sms_type') else '즉시전송'
        self.threaded_classification(
            self.classifier.classify_only,
            self.input_file,
            self.address_column,
            'dawn',
            dawn_type,
            sms_type,
            chasu
        )

    def start_classification_day_and_dawn(self):
        if not self.input_file or not self.address_column:
            messagebox.showwarning("입력 오류", "파일과 주소 컬럼을 선택하세요.")
            return
        if self.is_processing:
            return
        self.threaded_classification(self.classifier.classify_day_and_dawn, self.input_file, self.address_column)

    def update_zipcode(self, filetype):
        file = filedialog.askopenfilename(
            title="새 우편번호 csv 파일 선택",
            filetypes=[("CSV files", "*.csv")]
        )
        if file:
            self.classifier.update_zipcode_file(filetype, file)
            messagebox.showinfo("업데이트 완료", "우편번호 데이터가 교체되었습니다. 프로그램을 재시작하면 반영됩니다.")

    def create_chasu_tab(self):
        frame = self.tab_chasu
        for widget in frame.winfo_children():
            widget.destroy()
        tk.Label(frame, text='새벽배송 양식 파일을 업로드하고 저장합니다. 요청유형, 배송문자유형, 차수는 선택된 값으로 출력됩니다.', font=('맑은 고딕', 10), bg=self.bg_color).pack(pady=10)
        file_frame = tk.Frame(frame, bg=self.bg_color)
        file_frame.pack(pady=8)
        self.chasu_file_path = ''
        ttk.Button(file_frame, text='새벽배송 양식 파일 선택', command=self.select_chasu_file, width=22).pack(side=tk.LEFT)
        self.chasu_file_label = tk.Label(file_frame, text='(선택된 파일 없음)', fg='#1976d2', bg=self.bg_color, anchor='w', font=('맑은 고딕', 10))
        self.chasu_file_label.pack(side=tk.LEFT, padx=10)
        # 요청유형/배송문자유형 옵션 추가
        opt_row = tk.Frame(frame, bg=self.bg_color)
        opt_row.pack(anchor='w', pady=4)
        tk.Label(opt_row, text='요청유형:', bg=self.bg_color).pack(side=tk.LEFT, padx=2)
        self.chasu_dawn_type_var = tk.StringVar(value='배송대행')
        ttk.Combobox(opt_row, textvariable=self.chasu_dawn_type_var, values=['배송대행', '택배대행'], state='readonly', width=9).pack(side=tk.LEFT, padx=2)
        tk.Label(opt_row, text='배송문자유형:', bg=self.bg_color).pack(side=tk.LEFT, padx=2)
        self.chasu_sms_type_var = tk.StringVar(value='즉시전송')
        ttk.Combobox(opt_row, textvariable=self.chasu_sms_type_var, values=['즉시전송', '7시전송'], state='readonly', width=9).pack(side=tk.LEFT, padx=2)
        chasu_row = tk.Frame(frame, bg=self.bg_color)
        chasu_row.pack(anchor='w', pady=8)
        tk.Label(chasu_row, text='차수:', bg=self.bg_color).pack(side=tk.LEFT, padx=2)
        self.chasu_tab_var = tk.StringVar(value='차수없음')
        for chasu in ['차수없음', '1차', '2차', '3차', '4차', '5차', '6차', '7차', '8차', '9차', '10차']:
            ttk.Radiobutton(chasu_row, text=chasu, variable=self.chasu_tab_var, value=chasu).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame, text='유형 및 차수 반영저장', command=self.save_chasu_added_file, width=20).pack(pady=15)
        self.chasu_status = tk.Label(frame, text='', fg='green', bg=self.bg_color, font=('맑은 고딕', 10))
        self.chasu_status.pack(pady=5)

    def select_chasu_file(self):
        file = filedialog.askopenfilename(title='새벽배송 양식 파일 선택', filetypes=[('Excel files', '*.xlsx *.xls')])
        if file:
            self.chasu_file_path = file
            fname = os.path.basename(file)
            self.chasu_file_label.config(text=f'선택: {fname}')
        else:
            self.chasu_file_path = ''
            self.chasu_file_label.config(text='(선택된 파일 없음)')

    def save_chasu_added_file(self):
        import os
        import datetime
        import pandas as pd
        if not self.chasu_file_path:
            self.chasu_status.config(text='파일을 선택하세요.', fg='red')
            return
        chasu = self.chasu_tab_var.get()
        dawn_type = self.chasu_dawn_type_var.get() if hasattr(self, 'chasu_dawn_type_var') else '배송대행'
        sms_type = self.chasu_sms_type_var.get() if hasattr(self, 'chasu_sms_type_var') else '즉시전송'
        try:
            df = pd.read_excel(self.chasu_file_path, dtype=str)
            template_path = os.path.join('data', '새벽배송양식.xlsx')
            template_cols = list(pd.read_excel(template_path, nrows=0).columns)
            tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            def normalize(s):
                import re
                return re.sub(r'[^가-힣a-zA-Z0-9]', '', str(s)).lower()
            orig_cols = list(df.columns)
            orig_norm = {normalize(c): c for c in orig_cols}
            col_map = {}
            for tcol in template_cols:
                tnorm = normalize(tcol)
                if tcol in orig_cols:
                    col_map[tcol] = tcol
                    continue
                if tnorm in orig_norm:
                    col_map[tcol] = orig_norm[tnorm]
                    continue
                matches = [(len(oc), oc) for oc in orig_cols if tnorm in normalize(oc) or normalize(oc) in tnorm]
                if matches:
                    matches.sort()
                    col_map[tcol] = matches[0][1]
                    continue
                for key, cands in COLUMN_MAP.items():
                    for cand in cands:
                        if normalize(cand) == tnorm:
                            val = find_column(df.iloc[0], key)
                            if val != '':
                                col_map[tcol] = key
                            break
            result = []
            for _, row in df.iterrows():
                new_row = {col: '' for col in template_cols}
                for tcol in template_cols:
                    if tcol in col_map and col_map[tcol] in row:
                        val = row[col_map[tcol]]
                    elif tcol in col_map and col_map[tcol] in COLUMN_MAP:
                        val = find_column(row, col_map[tcol])
                    else:
                        val = ''
                    new_row[tcol] = val
                new_row['배송요청일*'] = tomorrow
                new_row['상품명*'] = DeliveryClassifier.make_dawn_product_name(row, chasu)
                road_addr = clean_address_for_output(replace_company_name(find_column(row, '수령자도로명주소')))
                detail_addr = find_column(row, '수령자상세주소')
                if detail_addr and str(detail_addr).strip() and str(detail_addr).strip().lower() != 'nan':
                    full_addr = f"{road_addr} {str(detail_addr).strip()}"
                else:
                    full_addr = road_addr
                new_row['수령자 도로명 주소*'] = full_addr.strip()
                new_row['수령자 상세 주소'] = ''
                new_row['수령자 연락처*'] = format_phone_number(find_column(row, '수령자연락처'))
                dawn_msg = find_column(row, '배송메시지')
                new_row['출입방법'] = dawn_msg
                def is_etc_condition(val):
                    if val is None:
                        return True
                    sval = str(val).strip()
                    if sval == '' or sval.lower() == 'nan' or sval == '0 / 0':
                        return True
                    return False
                for col in template_cols:
                    if '배송 받을 장소' in col:
                        new_row[col] = '기타' if is_etc_condition(dawn_msg) else '문 앞'
                    if '배송 받을 장소 상세' in col:
                        new_row[col] = '문앞 배송해주세요. 불가시 경비실호출 및 택배보관함배송'
                    if '배송 유형' in col:
                        new_row[col] = dawn_type or '배송대행'
                    if '배송 문자 전송 시점' in col:
                        new_row[col] = sms_type or '즉시전송'
                    if '출입방법' in col.replace(' ', ''):
                        new_row[col] = dawn_msg
                if '주문자 주소' in new_row:
                    new_row['주문자 주소'] = replace_company_name(find_column(row, '주문자주소'))
                if '우편번호' in new_row:
                    zipcode = find_column(row, '우편번호')
                    if zipcode:
                        try:
                            new_row['우편번호'] = str(int(float(zipcode))).zfill(5)
                        except:
                            new_row['우편번호'] = str(zipcode).zfill(5)
                result.append(new_row)
            df_result = pd.DataFrame(result, columns=template_cols)
            df_result = df_result.replace('nan', '').fillna('')
            for col in COLUMN_MAP['수령인연락처1']:
                if col in df_result.columns:
                    df_result[col] = df_result[col].apply(format_phone_number)
            for col in COLUMN_MAP['수령인연락처2']:
                if col in df_result.columns:
                    df_result[col] = df_result[col].apply(format_phone_number)
            for col in COLUMN_MAP['구매자연락처']:
                if col in df_result.columns:
                    df_result[col] = df_result[col].apply(format_phone_number)
            out_dir = os.path.dirname(self.chasu_file_path)
            base = os.path.splitext(os.path.basename(self.chasu_file_path))[0] + f'_{chasu}추가'
            out_path = self.classifier.get_incremented_filename(base, 'xlsx', out_dir)
            self.classifier.save_with_yellow_header(df_result, out_path, header=template_cols)
            self.chasu_status.config(text=f'저장 완료: {os.path.basename(out_path)}', fg='green')
            msg = f'저장 완료: {os.path.basename(out_path)} | 저장 파일: {out_path}'
            self.show_custom_message('완료', msg, file_path=out_path)
        except Exception as e:
            self.chasu_status.config(text=f'오류: {str(e)}', fg='red')

    def create_data_tab(self):
        """데이터 관리 탭 - 관리자 전용 data 폴더 관리"""
        frame = self.tab_data
        for widget in frame.winfo_children():
            widget.destroy()
            
        if not self.admin_authenticated:
            tk.Label(frame, text="관리자 인증이 필요합니다.", font=("맑은 고딕", 12)).pack(pady=20)
            return
            
        # 상단 제목
        tk.Label(frame, text="데이터 파일 관리 (관리자 전용)", 
                font=("맑은 고딕", 14, "bold"), fg=self.naver_green).pack(pady=15)
        
        # 파일 목록 표시 
        file_frame = ttk.LabelFrame(frame, text="Data 폴더 파일 목록", padding=10)
        file_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # 파일 목록을 표시할 Treeview
        columns = ("파일명", "크기", "수정일자")
        self.file_tree = ttk.Treeview(file_frame, columns=columns, show="headings")
        
        # 열 설정
        for col in columns:
            self.file_tree.heading(col, text=col)
        
        self.file_tree.column("파일명", width=250)
        self.file_tree.column("크기", width=100)
        self.file_tree.column("수정일자", width=150)
        
        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(file_frame, orient="vertical", command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=scrollbar.set)
        
        self.file_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 버튼 프레임
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=15, fill="x", padx=20)
        
        ttk.Button(btn_frame, text="새 파일 추가", 
                  command=self.add_data_file).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="선택 파일 수정", 
                  command=self.edit_data_file).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="파일 목록 새로고침", 
                  command=self.refresh_data_files).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="선택 파일 삭제", 
                  command=self.delete_data_file).pack(side="left", padx=5)
        
        # 파일 목록 로드
        self.refresh_data_files()
        
        # 도움말 메시지
        tk.Label(frame, text="※ 주의: data 폴더의 파일은 프로그램 동작에 중요합니다. 삭제 시 기능이 작동하지 않을 수 있습니다.", 
                fg="red", font=("맑은 고딕", 9)).pack(pady=10)
    
    def refresh_data_files(self):
        """Data 폴더의 파일 목록을 새로고침"""
        # 기존 목록 삭제
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
            
        # data 폴더 경로
        data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
        
        try:
            # 폴더가 없으면 생성
            if not os.path.exists(data_path):
                os.makedirs(data_path)
                
            # 파일 목록 가져오기
            files = os.listdir(data_path)
            for filename in files:
                file_path = os.path.join(data_path, filename)
                if os.path.isfile(file_path):
                    # 파일 정보 가져오기
                    file_size = os.path.getsize(file_path)
                    size_str = f"{file_size / 1024:.1f} KB" if file_size >= 1024 else f"{file_size} bytes"
                    
                    # 수정 시간
                    mod_time = os.path.getmtime(file_path)
                    mod_time_str = datetime.datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d %H:%M")
                    
                    # 트리뷰에 추가
                    self.file_tree.insert("", "end", values=(filename, size_str, mod_time_str))
        except Exception as e:
            messagebox.showerror("오류", f"파일 목록을 불러오는 중 오류가 발생했습니다.\n{str(e)}")
    
    def add_data_file(self):
        """새 데이터 파일 추가"""
        file_path = filedialog.askopenfilename(
            title="추가할 파일 선택",
            filetypes=[
                ("모든 파일", "*.*"),
                ("CSV 파일", "*.csv"),
                ("엑셀 파일", "*.xlsx"),
                ("텍스트 파일", "*.txt")
            ]
        )
        
        if file_path:
            try:
                # data 폴더 경로
                data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
                
                # 파일 이름 가져오기
                filename = os.path.basename(file_path)
                
                # 목적지 경로
                dest_path = os.path.join(data_path, filename)
                
                # 이미 파일이 있는지 확인
                if os.path.exists(dest_path):
                    overwrite = messagebox.askyesno(
                        "파일 덮어쓰기",
                        f"'{filename}' 파일이 이미 존재합니다. 덮어쓰시겠습니까?"
                    )
                    if not overwrite:
                        return
                
                # 파일 복사
                import shutil
                shutil.copy2(file_path, dest_path)
                
                messagebox.showinfo("완료", f"'{filename}' 파일이 data 폴더에 추가되었습니다.")
                
                # 파일 목록 새로고침
                self.refresh_data_files()
                
            except Exception as e:
                messagebox.showerror("오류", f"파일 추가 중 오류가 발생했습니다.\n{str(e)}")
    
    def edit_data_file(self):
        """선택한 데이터 파일 편집"""
        selected = self.file_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "편집할 파일을 선택하세요.")
            return
            
        # 선택한 파일 이름 가져오기
        filename = self.file_tree.item(selected[0])['values'][0]
        
        # 파일 경로
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', filename)
        
        # 파일 확장자 확인
        ext = os.path.splitext(filename)[1].lower()
        
        try:
            # 파일 종류별 처리
            if ext == '.csv':
                # CSV 파일은 메모장으로 열기
                import subprocess
                subprocess.Popen(['notepad.exe', file_path])
            elif ext == '.xlsx':
                # Excel 파일은 시스템 기본 앱으로 열기
                os.startfile(file_path)
            elif ext == '.txt':
                # 텍스트 파일은 메모장으로 열기
                import subprocess
                subprocess.Popen(['notepad.exe', file_path])
            else:
                # 기타 파일은 시스템 기본 앱으로 열기
                os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("오류", f"파일 편집 중 오류가 발생했습니다.\n{str(e)}")
    
    def delete_data_file(self):
        """선택한 데이터 파일 삭제"""
        selected = self.file_tree.selection()
        if not selected:
            messagebox.showwarning("선택 필요", "삭제할 파일을 선택하세요.")
            return
            
        # 선택한 파일 이름 가져오기
        filename = self.file_tree.item(selected[0])['values'][0]
        
        # 필수 파일인지 확인 (새벽배송양식.xlsx, api_key.txt, 당일_우편번호.csv, 새벽_우편번호.csv)
        essential_files = ['새벽배송양식.xlsx', 'api_key.txt', '당일_우편번호.csv', '새벽_우편번호.csv']
        if filename in essential_files:
            confirm = messagebox.askyesno(
                "필수 파일 삭제 경고",
                f"'{filename}'은(는) 프로그램 동작에 필수적인 파일입니다.\n정말로 삭제하시겠습니까? 삭제 시 프로그램이 정상 작동하지 않을 수 있습니다.",
                icon='warning'
            )
        else:
            confirm = messagebox.askyesno(
                "파일 삭제 확인", 
                f"'{filename}' 파일을 정말로 삭제하시겠습니까?"
            )
            
        if confirm:
            try:
                # 파일 경로
                file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', filename)
                
                # 파일 삭제
                os.remove(file_path)
                
                messagebox.showinfo("완료", f"'{filename}' 파일이 삭제되었습니다.")
                
                # 파일 목록 새로고침
                self.refresh_data_files()
                
            except Exception as e:
                messagebox.showerror("오류", f"파일 삭제 중 오류가 발생했습니다.\n{str(e)}")

    def start_classification_all_in_one(self):
        if not self.input_file or not self.address_column:
            messagebox.showwarning("입력 오류", "파일과 주소 컬럼을 선택하세요.")
            return
        if self.is_processing:
            return
        self.threaded_classification(self._run_all_in_one)

    def _run_all_in_one(self):
        try:
            print("분류 시작")
            df = read_excel_file(self.input_file)
            def progress_callback(progress):
                self.root.after(0, lambda: self.update_progress(progress))
            result_df = self.classifier.classify_all_delivery_types(df, input_file_path=self.input_file, progress_callback=progress_callback)
            # os.startfile(os.path.dirname(self.input_file))  # 폴더 미리 열기 코드 삭제
            return getattr(self.classifier, 'last_output_path', None)
        except Exception as e:
            import traceback
            messagebox.showerror("오류", f"분류 중 오류 발생:\n{str(e)}\n{traceback.format_exc()}")
            self.progress_label.config(text=f'오류: {str(e)}', fg='red')

    def create_invoice_tab(self):
        frame = self.tab_invoice
        for widget in frame.winfo_children():
            widget.destroy()
        # 안내문구
        tk.Label(frame, text='허우적송장, 롯데송장, 컬리송장번호를 자동으로 원데이터에 매칭합니다. 미매칭시 미매칭 데이터가 출력됩니다.', font=('맑은 고딕', 10), bg=self.bg_color).pack(pady=10)
        # 파일 선택 그룹
        file_group = ttk.Labelframe(frame, text="송장 매칭 파일 선택 (배송분류탭에서 구분처리한양식 + 허우적본사 송장출력양식 + 컬리송장출력양식)", padding=10, style='TLabelframe')
        file_group.pack(pady=8, fill="x", padx=35)
        file_frame = tk.Frame(file_group, bg=self.bg_color)
        file_frame.pack(pady=4)
        self.invoice_order_file = ''
        self.invoice_day_file = ''
        self.invoice_dawn_file = ''
        ttk.Button(file_frame, text='구분된 주문양식 선택', command=self.select_invoice_order_file, width=22).pack(side=tk.LEFT, padx=4)
        self.invoice_order_label = tk.Label(file_frame, text='(선택 없음)', fg='#1976d2', bg=self.bg_color, anchor='w', font=('맑은 고딕', 10))
        self.invoice_order_label.pack(side=tk.LEFT, padx=10)
        file_frame2 = tk.Frame(file_group, bg=self.bg_color)
        file_frame2.pack(pady=4)
        ttk.Button(file_frame2, text='당일&택배 송장 선택', command=self.select_invoice_day_file, width=22).pack(side=tk.LEFT, padx=4)
        self.invoice_day_label = tk.Label(file_frame2, text='(선택 없음)', fg='#1976d2', bg=self.bg_color, anchor='w', font=('맑은 고딕', 10))
        self.invoice_day_label.pack(side=tk.LEFT, padx=10)
        file_frame3 = tk.Frame(file_group, bg=self.bg_color)
        file_frame3.pack(pady=4)
        ttk.Button(file_frame3, text='새벽 송장 선택', command=self.select_invoice_dawn_file, width=22).pack(side=tk.LEFT, padx=4)
        self.invoice_dawn_label = tk.Label(file_frame3, text='(선택 없음)', fg='#1976d2', bg=self.bg_color, anchor='w', font=('맑은 고딕', 10))
        self.invoice_dawn_label.pack(side=tk.LEFT, padx=10)
        # 실행 버튼
        ttk.Button(file_group, text='송장 매칭 실행', command=self.start_invoice_matching, width=24).pack(pady=10)
        self.invoice_status = tk.Label(file_group, text='', fg='green', bg=self.bg_color, font=('맑은 고딕', 10))
        self.invoice_status.pack(pady=5)
        self.invoice_result = tk.Label(file_group, text='', fg='#1976d2', bg=self.bg_color, font=('맑은 고딕', 10))
        self.invoice_result.pack(pady=5)
        # 구분선
        sep = tk.Frame(frame, height=2, bd=1, relief='sunken', bg='#e0e0e0')
        sep.pack(fill='x', padx=10, pady=18)
        # 새벽배송 송장매칭 그룹
        dawn_group = ttk.Labelframe(frame, text="새벽배송 전용 송장매칭 (새벽배송양식 + 새벽송장번호양식)", padding=10, style='TLabelframe')
        dawn_group.pack(pady=8, fill="x", padx=35)
        dawn_file_frame = tk.Frame(dawn_group, bg=self.bg_color)
        dawn_file_frame.pack(pady=4)
        self.dawn_invoice_order_file = ''
        self.dawn_invoice_file = ''
        ttk.Button(dawn_file_frame, text='새벽배송양식 선택', command=self.select_dawn_invoice_order_file, width=22).pack(side=tk.LEFT, padx=4)
        self.dawn_invoice_order_label = tk.Label(dawn_file_frame, text='(선택 없음)', fg='#1976d2', bg=self.bg_color, anchor='w', font=('맑은 고딕', 10))
        self.dawn_invoice_order_label.pack(side=tk.LEFT, padx=10)
        dawn_file_frame2 = tk.Frame(dawn_group, bg=self.bg_color)
        dawn_file_frame2.pack(pady=4)
        ttk.Button(dawn_file_frame2, text='새벽송장번호양식 선택', command=self.select_dawn_invoice_file, width=22).pack(side=tk.LEFT, padx=4)
        self.dawn_invoice_label = tk.Label(dawn_file_frame2, text='(선택 없음)', fg='#1976d2', bg=self.bg_color, anchor='w', font=('맑은 고딕', 10))
        self.dawn_invoice_label.pack(side=tk.LEFT, padx=10)
        ttk.Button(dawn_group, text='새벽 송장매칭 실행', command=self.start_dawn_invoice_matching, width=24).pack(pady=10)
        self.dawn_invoice_status = tk.Label(dawn_group, text='', fg='green', bg=self.bg_color, font=('맑은 고딕', 10))
        self.dawn_invoice_status.pack(pady=5)
        self.dawn_invoice_result = tk.Label(dawn_group, text='', fg='#1976d2', bg=self.bg_color, font=('맑은 고딕', 10))
        self.dawn_invoice_result.pack(pady=5)

    def select_dawn_invoice_order_file(self):
        file = filedialog.askopenfilename(title='새벽배송양식 선택', filetypes=[('Excel files', '*.xlsx *.xls')])
        if file:
            self.dawn_invoice_order_file = file
            fname = os.path.basename(file)
            self.dawn_invoice_order_label.config(text=f'선택: {fname}')
        else:
            self.dawn_invoice_order_file = ''
            self.dawn_invoice_order_label.config(text='(선택 없음)')

    def select_dawn_invoice_file(self):
        file = filedialog.askopenfilename(title='새벽송장번호양식 선택', filetypes=[('Excel files', '*.xlsx *.xls')])
        if file:
            self.dawn_invoice_file = file
            fname = os.path.basename(file)
            self.dawn_invoice_label.config(text=f'선택: {fname}')
        else:
            self.dawn_invoice_file = ''
            self.dawn_invoice_label.config(text='(선택 없음)')

    def start_dawn_invoice_matching(self):
        import pandas as pd
        import os
        import difflib
        import re
        try:
            if not self.dawn_invoice_order_file:
                self.dawn_invoice_status.config(text='새벽배송양식을 업로드하세요.', fg='red')
                return
            if not self.dawn_invoice_file:
                self.dawn_invoice_status.config(text='새벽송장번호양식을 업로드하세요.', fg='red')
                return
            order_df = pd.read_excel(self.dawn_invoice_order_file)
            invoice_df = pd.read_excel(self.dawn_invoice_file)
            # 컬럼 자동 매핑
            def auto_map_col(df, std_names, force_addr=False):
                if force_addr:
                    addr_cols = [col for col in df.columns if '주소' in str(col)]
                    for c in std_names:
                        for col in addr_cols:
                            if c == col:
                                return col
                    def normalize(s):
                        return re.sub(r'[^-가-힣]', '', str(s)).lower()
                    norm_std = [normalize(c) for c in std_names]
                    for col in addr_cols:
                        ncol = normalize(col)
                        for nstd, std in zip(norm_std, std_names):
                            if nstd in ncol or ncol in nstd:
                                return col
                    if addr_cols:
                        return addr_cols[0]
                for c in std_names:
                    if c in df.columns:
                        return c
                def normalize(s):
                    return re.sub(r'[^-가-힣]', '', str(s)).lower()
                norm_std = [normalize(c) for c in std_names]
                for col in df.columns:
                    ncol = normalize(col)
                    for nstd, std in zip(norm_std, std_names):
                        if nstd in ncol or ncol in nstd:
                            return col
                matches = difflib.get_close_matches(std_names[0], df.columns, n=1, cutoff=0.7)
                if matches:
                    return matches[0]
                return None
            def clean_address_for_matching(address):
                if pd.isna(address):
                    return ''
                s = str(address).lower()
                s = re.sub(r'\([^)]*\)', '', s)
                s = re.sub(r'\[[^\]]*\]', '', s)
                s = re.sub(r'\s+', ' ', s)
                s = re.sub(r'[^\w가-힣\s\-\.]', '', s)
                return s.strip()
            def calculate_similarity(s1, s2):
                from difflib import SequenceMatcher
                if pd.isna(s1) or pd.isna(s2):
                    return 0
                return SequenceMatcher(None, str(s1), str(s2)).ratio()
            # 매핑
            order_name_col = auto_map_col(order_df, ['받는분', '수령인', '수취인', '수령자'])
            order_addr_col = auto_map_col(order_df, ['받는분주소', '수령인주소', '주소', '수령자 도로명 주소', '수령자 주소'])
            order_prod_col = auto_map_col(order_df, ['상품주문번호', '주문번호', '거래처주문코드', '거래처 주문번호'])
            invoice_name_col = auto_map_col(invoice_df, ['받는분', '수령인', '수취인', '수령자'])
            invoice_addr_col = auto_map_col(invoice_df, ['받는분주소', '수령인주소', '주소', '수령자 도로명 주소', '수령자 주소'])
            invoice_prod_col = auto_map_col(invoice_df, ['상품주문번호', '주문번호', '거래처주문코드', '거래처 주문번호'])
            invoice_num_col = auto_map_col(invoice_df, ['운송장번호', '송장번호'])
            if not all([order_name_col, order_addr_col, invoice_name_col, invoice_addr_col, invoice_num_col]):
                self.dawn_invoice_status.config(text='필수 컬럼을 찾을 수 없습니다.', fg='red')
                return
            order_df['__clean_name'] = order_df[order_name_col].apply(clean_address_for_matching)
            order_df['__clean_addr'] = order_df[order_addr_col].apply(clean_address_for_matching)
            invoice_df['__clean_name'] = invoice_df[invoice_name_col].apply(clean_address_for_matching)
            invoice_df['__clean_addr'] = invoice_df[invoice_addr_col].apply(clean_address_for_matching)
            invoice_df['matched'] = False
            order_df['송장번호'] = ''
            order_df['배송사명'] = ''
            matched_count = 0
            for idx, row in order_df.iterrows():
                name = row['__clean_name']
                addr = row['__clean_addr']
                prod_code = row[order_prod_col] if order_prod_col else None
                # 1. 상품주문번호로 우선 매칭
                match = None
                if order_prod_col and invoice_prod_col and pd.notna(prod_code):
                    match_df = invoice_df[(invoice_df[invoice_prod_col] == prod_code) & (invoice_df['matched'] == False)]
                    if not match_df.empty:
                        match = match_df.iloc[0]
                # 2. 이름+주소 유사도 매칭
                if match is None:
                    candidates = invoice_df[invoice_df['matched'] == False].copy()
                    candidates['name_similarity'] = candidates['__clean_name'].apply(lambda x: calculate_similarity(name, x))
                    candidates['addr_similarity'] = candidates['__clean_addr'].apply(lambda x: calculate_similarity(addr, x))
                    matches = candidates[(candidates['name_similarity'] >= 0.8) & (candidates['addr_similarity'] >= 0.8)]
                    if not matches.empty:
                        best_idx = matches['name_similarity'].idxmax()
                        match = invoice_df.loc[best_idx]
                if match is not None:
                    order_df.at[idx, '송장번호'] = str(match[invoice_num_col])
                    order_df.at[idx, '배송사명'] = '컬리넥스트마일'
                    invoice_df.at[match.name, 'matched'] = True
                    matched_count += 1
            total = len(order_df)
            unmatched = total - matched_count
            match_rate = (matched_count / total * 100) if total > 0 else 0
            out_dir = os.path.dirname(self.dawn_invoice_order_file)
            base = os.path.splitext(os.path.basename(self.dawn_invoice_order_file))[0] + '_새벽송장매칭'
            out_path = self.classifier.get_incremented_filename(base, 'xlsx', out_dir)
            # 임시 컬럼 삭제
            order_df_clean = order_df.drop(columns=['__clean_name', '__clean_addr'])
            # --- 컬럼 순서: 배송사명 -> 송장번호 순서로 재배치 ---
            cols = list(order_df_clean.columns)
            if '배송사명' in cols and '송장번호' in cols:
                cols.remove('배송사명')
                cols.remove('송장번호')
                # 배송사명, 송장번호를 원하는 위치에 삽입 (맨 뒤에)
                cols.append('배송사명')
                cols.append('송장번호')
                order_df_clean = order_df_clean[cols]
            # --- 미매칭/매칭 데이터 분리 저장 ---
            matched_df = order_df_clean[(order_df_clean['송장번호'].notna()) & (order_df_clean['송장번호'] != '')]
            unmatched_df = order_df_clean[(order_df_clean['송장번호'].isna()) | (order_df_clean['송장번호'] == '')]
            for df in [matched_df, unmatched_df]:
                if '우편번호' in df.columns:
                    def clean_zipcode(val):
                        if pd.isna(val) or str(val).strip() in ['', 'nan', 'None']:
                            return ''
                        try:
                            return str(int(float(val))).zfill(5)
                        except:
                            return str(val).zfill(5)
                    df['우편번호'] = df['우편번호'].apply(clean_zipcode)
            # 미매칭 사유 컬럼 추가
            if not unmatched_df.empty:
                unmatched_df = unmatched_df.copy()
                unmatched_df['미매칭사유'] = '매칭실패'
            self.classifier.save_with_yellow_header(matched_df, out_path)
            if not unmatched_df.empty:
                unmatched_path = self.classifier.get_incremented_filename(base + '_미매칭', 'xlsx', out_dir)
                self.classifier.save_with_yellow_header(unmatched_df, unmatched_path)
            result_msg = f"총 주문: {total}건 | 매칭: {matched_count}건 | 미매칭: {unmatched}건\n매칭률: {match_rate:.2f}%"
            self.dawn_invoice_result.config(text=result_msg)
            self.dawn_invoice_status.config(text=f'저장 완료: {os.path.basename(out_path)}', fg='green')
            self.root.after(0, lambda: self.show_custom_message('완료', f'저장 완료: {os.path.basename(out_path)} | 저장 파일: {out_path}', file_path=out_path))
        except Exception as e:
            import traceback
            self.dawn_invoice_status.config(text=f'오류: {str(e)}', fg='red')
            import tkinter.messagebox as mb
            mb.showerror('오류', f'매칭 중 오류 발생:\n{str(e)}\n{traceback.format_exc()}')

    def select_invoice_order_file(self):
        file = filedialog.askopenfilename(title='구분된 주문양식 선택', filetypes=[('Excel files', '*.xlsx *.xls')])
        if file:
            self.invoice_order_file = file
            fname = os.path.basename(file)
            self.invoice_order_label.config(text=f'선택: {fname}')
        else:
            self.invoice_order_file = ''
            self.invoice_order_label.config(text='(선택 없음)')
    def select_invoice_day_file(self):
        file = filedialog.askopenfilename(title='당일&택배 송장 선택', filetypes=[('Excel files', '*.xlsx *.xls')])
        if file:
            self.invoice_day_file = file
            fname = os.path.basename(file)
            self.invoice_day_label.config(text=f'선택: {fname}')
        else:
            self.invoice_day_file = ''
            self.invoice_day_label.config(text='(선택 없음)')
    def select_invoice_dawn_file(self):
        file = filedialog.askopenfilename(title='새벽 송장 선택', filetypes=[('Excel files', '*.xlsx *.xls')])
        if file:
            self.invoice_dawn_file = file
            fname = os.path.basename(file)
            self.invoice_dawn_label.config(text=f'선택: {fname}')
        else:
            self.invoice_dawn_file = ''
            self.invoice_dawn_label.config(text='(선택 없음)')
    def start_invoice_matching(self):
        import pandas as pd
        import os
        import difflib
        import re
        import logging
        
        def auto_map_col(df, std_names, force_addr=False):
            # 주소 매핑일 경우 '주소'가 포함된 컬럼만 우선 매핑
            if force_addr:
                addr_cols = [col for col in df.columns if '주소' in str(col)]
                for c in std_names:
                    for col in addr_cols:
                        if c == col:
                            return col
                # 부분일치
                def normalize(s):
                    import re
                    return re.sub(r'[^\w]', '', str(s)).lower()
                norm_std = [normalize(c) for c in std_names]
                for col in addr_cols:
                    ncol = normalize(col)
                    for nstd, std in zip(norm_std, std_names):
                        if nstd in ncol or ncol in nstd:
                            return col
                if addr_cols:
                    return addr_cols[0]  # fallback: 주소 포함된 첫 컬럼
            # 기존 로직
            for c in std_names:
                if c in df.columns:
                    return c
            def normalize(s):
                import re
                return re.sub(r'[^\w]', '', str(s)).lower()
            norm_std = [normalize(c) for c in std_names]
            for col in df.columns:
                ncol = normalize(col)
                for nstd, std in zip(norm_std, std_names):
                    if nstd in ncol or ncol in nstd:
                        return col
            matches = difflib.get_close_matches(std_names[0], df.columns, n=1, cutoff=0.7)
            if matches:
                return matches[0]
            return None

        def strong_clean(val):
            if pd.isna(val):
                return ''
            s = str(val).lower()
            s = re.sub(r'\[.*?\]', '', s)
            s = re.sub(r'\(.*?\)', '', s)
            s = re.sub(r'\[파일집명!.*?\]', '', s)
            s = re.sub(r'[^\w가-힣]', '', s)
            return s.strip()

        def clean_address_for_matching(address):
            """매칭을 위한 주소 정규화 함수"""
            if pd.isna(address):
                return ''
            s = str(address).lower()
            # 괄호 내용 제거
            s = re.sub(r'\([^)]*\)', '', s)
            s = re.sub(r'\[[^\]]*\]', '', s)
            # 공백 정규화
            s = re.sub(r'\s+', ' ', s)
            # 특수문자 중 일부만 제거 (주소 구분에 중요한 문자는 보존)
            s = re.sub(r'[^\w가-힣\s\-\.]', '', s)
            return s.strip()

        def calculate_similarity(s1, s2):
            """두 문자열의 유사도 계산"""
            from difflib import SequenceMatcher
            if pd.isna(s1) or pd.isna(s2):
                return 0
            return SequenceMatcher(None, str(s1), str(s2)).ratio()

        def find_best_match(order_row, invoice_df, threshold=0.8, invoice_prod_col=None, order_prod_col=None, code_col=None, code_filter=None):
            """가장 유사한 송장 정보 찾기 (상품주문번호까지 포함, 1:1 매칭)"""
            name = order_row['__clean_name']
            addr = order_row['__clean_addr']
            order_prod_code = order_row.get(order_prod_col, None) if order_prod_col else None
            print(f"[DEBUG] 매칭 시도: 주문번호({order_prod_col})={order_prod_code}")
            # 1. 집하지점코드 등 필터 적용
            candidates = invoice_df[invoice_df['matched'] == False]
            if code_col and code_filter is not None:
                candidates = candidates[candidates[code_col].astype(str).str.startswith(str(code_filter))]
            # 2. 상품주문번호까지 포함해서 후보 필터링
            if order_prod_code and invoice_prod_col and invoice_prod_col in candidates.columns:
                candidates = candidates[candidates[invoice_prod_col] == order_prod_code]
                print(f"[DEBUG] 후보 송장 수: {len(candidates)} (송장번호 컬럼: {invoice_prod_col})")
            else:
                print(f"[DEBUG] 상품주문번호 기준 후보 없음, 전체 후보 수: {len(candidates)}")
            # 3. 이름/주소 유사도 계산
            if not candidates.empty:
                candidates = candidates.copy()
                candidates['name_similarity'] = candidates['__clean_name'].apply(lambda x: calculate_similarity(name, x))
                candidates['addr_similarity'] = candidates['__clean_addr'].apply(lambda x: calculate_similarity(addr, x))
                # 4. 유사도 기준으로 필터링
                matches = candidates[(candidates['name_similarity'] >= threshold) & (candidates['addr_similarity'] >= threshold)]
                print(f"[DEBUG] 유사도 기준 통과 후보 수: {len(matches)}")
                if not matches.empty:
                    # 5. 가장 높은 유사도 가진 행 선택 (1:1 매칭)
                    best_idx = matches['name_similarity'].idxmax()  # 이 인덱스는 invoice_df의 인덱스임
                    best_match = invoice_df.loc[best_idx]  # 반드시 원본에서 가져오기
                    invoice_df.at[best_idx, 'matched'] = True  # 원본에 True로!
                    print(f"[DEBUG] 매칭 성공: 송장번호={best_match.get('송장번호', None)}")
                    return best_match, {
                        'name_similarity': best_match['name_similarity'] if 'name_similarity' in best_match else 1.0,
                        'addr_similarity': best_match['addr_similarity'] if 'addr_similarity' in best_match else 1.0
                    }
            print(f"[DEBUG] 매칭 실패")
            return None, None

        def log_matching_details(order_row, invoice_row, similarity_scores, matched=False):
            """매칭 상세 정보 로깅"""
            log = {
                'order_name': order_row['__clean_name'],
                'order_addr': order_row['__clean_addr'],
                'invoice_name': invoice_row['__clean_name'] if invoice_row is not None else None,
                'invoice_addr': invoice_row['__clean_addr'] if invoice_row is not None else None,
                'name_similarity': similarity_scores['name_similarity'] if similarity_scores else None,
                'addr_similarity': similarity_scores['addr_similarity'] if similarity_scores else None,
                'matched': matched
            }
            print(f"[매칭상세] {log}")

        def analyze_matching_results(order_df, invoice_df, unmatched_rows):
            """매칭 결과 분석"""
            analysis = {
                'total_orders': len(order_df),
                'matched_orders': len(order_df) - len(unmatched_rows),
                'unmatched_orders': len(unmatched_rows),
                'match_rate': (len(order_df) - len(unmatched_rows)) / len(order_df) * 100 if len(order_df) > 0 else 0,
                'common_mismatch_patterns': []
            }
            
            # 미매칭 패턴 분석
            for row in unmatched_rows:
                pattern = {
                    'name_length': len(str(row['__clean_name'])),
                    'addr_length': len(str(row['__clean_addr'])),
                    'special_chars': bool(re.search(r'[^\w가-힣\s]', str(row['__clean_addr'])))
                }
                analysis['common_mismatch_patterns'].append(pattern)
            
            print(f"[매칭분석] 총 주문: {analysis['total_orders']}, 매칭: {analysis['matched_orders']}, "
                  f"미매칭: {analysis['unmatched_orders']}, 매칭률: {analysis['match_rate']:.2f}%")
            
            return analysis

        try:
            if not self.invoice_order_file:
                self.invoice_status.config(text='분류된 주문양식을 업로드하세요.', fg='red')
                return
            
            order_df = pd.read_excel(self.invoice_order_file)
            if self.invoice_day_file:
                day_df = pd.read_excel(self.invoice_day_file)
                day_name_col = auto_map_col(day_df, ['받는분', '수령인', '수취인'])
                day_addr_col = auto_map_col(day_df, ['받는분주소', '수령인주소', '주소'], force_addr=True)
                day_code_col = auto_map_col(day_df, ['배송점', '집하지점코드'])
                day_invoice_col = auto_map_col(day_df, ['송장번호', '운송장번호'])
                
                if not all([day_name_col, day_addr_col, day_code_col, day_invoice_col]):
                    raise ValueError(f"당일&택배 송장 파일의 실제 컬럼명: {list(day_df.columns)}")
                    
                day_df['__clean_name'] = day_df[day_name_col].apply(clean_address_for_matching)
                day_df['__clean_addr'] = day_df[day_addr_col].apply(clean_address_for_matching)
                day_df['matched'] = False
            else:
                day_df = None
                
            if self.invoice_dawn_file:
                dawn_df = pd.read_excel(self.invoice_dawn_file)
                dawn_name_col = auto_map_col(dawn_df, ['받는분', '수령인', '수취인', '수령자'])
                dawn_addr_col = auto_map_col(dawn_df, ['받는분주소', '수령인주소', '주소', '수령자 주소'], force_addr=True)
                dawn_invoice_col = auto_map_col(dawn_df, ['운송장번호', '송장번호'])
                
                if not all([dawn_name_col, dawn_addr_col, dawn_invoice_col]):
                    raise ValueError(f"새벽 송장 파일의 실제 컬럼명: {list(dawn_df.columns)}")
                    
                dawn_df['__clean_name'] = dawn_df[dawn_name_col].apply(clean_address_for_matching)
                dawn_df['__clean_addr'] = dawn_df[dawn_addr_col].apply(clean_address_for_matching)
                dawn_df['matched'] = False
            else:
                dawn_df = None
                
            order_name_col = auto_map_col(order_df, ['받는분', '수령인', '수취인'])
            order_addr_col = auto_map_col(order_df, ['받는분주소', '수령인주소', '주소'])
            order_code_col = auto_map_col(order_df, ['집하지점코드'])
            order_prod_col = auto_map_col(order_df, ['상품주문번호', '주문번호', '거래처주문코드'])
            print(f"[DEBUG] 주문 데이터 상품주문번호 컬럼: {order_prod_col}")
            
            if not all([order_name_col, order_addr_col, order_code_col]):
                raise ValueError(f"분류된 주문양식 파일의 실제 컬럼명: {list(order_df.columns)}")
                
            order_df['배송사명'] = ''
            order_df['송장번호'] = ''
            order_df['__clean_name'] = order_df[order_name_col].apply(clean_address_for_matching)
            order_df['__clean_addr'] = order_df[order_addr_col].apply(clean_address_for_matching)
            
            unmatched_rows = []
            for idx, row in order_df.iterrows():
                address = row['__clean_name']
                addr = row['__clean_addr']
                code = str(row[order_code_col]).strip() if pd.notna(row[order_code_col]) else ''
                
                try:
                    code_int = int(float(code)) if code else None
                    code_str = str(code_int) if code_int is not None else code
                except:
                    code_int = None
                    code_str = code
                    
                matched = False
                best_match = None
                similarity_scores = None
                
                # 당일/택배 매칭
                if day_df is not None:
                    # 상품주문번호 컬럼 찾기
                    day_prod_col = auto_map_col(day_df, ['상품주문번호', '주문번호', '거래처주문코드', '거래처 주문번호', '거래처주문번호'])
                    print(f"[DEBUG] 송장 데이터 상품주문번호 컬럼: {day_prod_col}")
                    
                    if code == '' or code == 'nan':
                        # 당일 매칭 (집하지점코드가 105, 108이 아닌 것)
                        best_match, similarity_scores = find_best_match(
                            row,
                            day_df,
                            threshold=0.8,
                            invoice_prod_col=day_prod_col,
                            order_prod_col=order_prod_col,
                            code_col=day_code_col,
                            code_filter=''  # ''로 시작하는 집하지점코드(당일)
                        )
                        if best_match is not None:
                            order_df.at[idx, '배송사명'] = '우리한방택배'
                            order_df.at[idx, '송장번호'] = str(best_match[day_invoice_col])
                            matched = True
                            log_matching_details(row, best_match, similarity_scores, matched)
                    elif code_str == '105':
                        # 택배 매칭
                        best_match, similarity_scores = find_best_match(
                            row,
                            day_df[day_df[day_code_col].astype(str).str.startswith('105')],
                            threshold=0.8,
                            invoice_prod_col=day_prod_col,
                            order_prod_col=order_prod_col
                        )
                        if best_match is not None:
                            order_df.at[idx, '배송사명'] = '롯데택배'
                            order_df.at[idx, '송장번호'] = str(best_match[day_invoice_col])
                            matched = True
                            log_matching_details(row, best_match, similarity_scores, matched)
                        
                # 새벽 매칭
                if not matched and code_str == '108' and dawn_df is not None:
                    best_match, similarity_scores = find_best_match(row, dawn_df, threshold=0.8)
                    if best_match is not None:
                        order_df.at[idx, '배송사명'] = '컬리넥스트마일'
                        order_df.at[idx, '송장번호'] = str(best_match[dawn_invoice_col])
                        matched = True
                        log_matching_details(row, best_match, similarity_scores, matched)
                        
                if not matched:
                    print(f"[미매칭] idx={idx}, name={address}, addr={addr}, code={code}")
                    unmatched_rows.append(row)
                    
            # 매칭 결과 분석
            analysis = analyze_matching_results(order_df, day_df if day_df is not None else dawn_df, unmatched_rows)
            
            # 매칭 결과 메시지 업데이트
            result_msg = f"총 주문: {analysis['total_orders']}건 | 매칭: {analysis['matched_orders']}건 | 미매칭: {analysis['unmatched_orders']}건\n매칭률: {analysis['match_rate']:.2f}%"
            self.invoice_result.config(text=result_msg)
            
            # 결과 저장
            out_dir = os.path.dirname(self.invoice_order_file)
            base = os.path.splitext(os.path.basename(self.invoice_order_file))[0] + '_송장매칭'
            out_path = self.classifier.get_incremented_filename(base, 'xlsx', out_dir)
            # 임시 컬럼 정확히 삭제 (drop 반환값을 새로운 변수에 할당)
            order_df_clean = order_df
            for col in ['__clean_name', '__clean_addr']:
                if col in order_df_clean.columns:
                    order_df_clean = order_df_clean.drop(columns=[col])
            # --- 매칭/미매칭 분리 및 우편번호 5자리 문자열 변환 ---
            matched_df = order_df_clean[(order_df_clean['송장번호'].notna()) & (order_df_clean['송장번호'] != '')]
            unmatched_df = order_df_clean[(order_df_clean['송장번호'].isna()) | (order_df_clean['송장번호'] == '')]
            for df in [matched_df, unmatched_df]:
                if '우편번호' in df.columns:
                    df['우편번호'] = df['우편번호'].astype(str).str.zfill(5)
            # 미매칭 사유 컬럼 추가
            if not unmatched_df.empty:
                unmatched_df = unmatched_df.copy()
                unmatched_df['미매칭사유'] = '매칭실패'
            self.classifier.save_with_yellow_header(matched_df, out_path)
            if not unmatched_df.empty:
                unmatched_path = self.classifier.get_incremented_filename(base + '_미매칭', 'xlsx', out_dir)
                self.classifier.save_with_yellow_header(unmatched_df, unmatched_path)
            
            # 매칭 결과 메시지 업데이트
            result_msg = f"총 주문: {analysis['total_orders']}건 | 매칭: {analysis['matched_orders']}건 | 미매칭: {analysis['unmatched_orders']}건\n매칭률: {analysis['match_rate']:.2f}%"
            self.invoice_result.config(text=result_msg)
            
            # 완료 메시지 표시
            msg = "송장 매칭이 완료되었습니다!"
            if out_path:
                msg += f" | 저장 파일: {out_path}"
            self.root.after(0, lambda: self.show_custom_message("완료", msg, file_path=out_path))
            
        except Exception as e:
            import traceback
            # 예외 발생 시에도 미매칭 데이터 저장 시도
            try:
                if 'order_df_clean' in locals():
                    unmatched_df = order_df_clean[(order_df_clean['송장번호'].isna()) | (order_df_clean['송장번호'] == '')]
                    if not unmatched_df.empty:
                        unmatched_df = unmatched_df.copy()
                        unmatched_df['미매칭사유'] = '매칭실패(예외발생)'
                        out_dir = os.path.dirname(self.invoice_order_file)
                        base = os.path.splitext(os.path.basename(self.invoice_order_file))[0] + '_송장매칭'
                        unmatched_path = self.classifier.get_incremented_filename(base + '_미매칭', 'xlsx', out_dir)
                        self.classifier.save_with_yellow_header(unmatched_df, unmatched_path)
            except Exception as e2:
                print('미매칭 데이터 저장 중 추가 오류:', e2)
            self.invoice_status.config(text=f'오류: {str(e)}', fg='red')
            import tkinter.messagebox as mb
            mb.showerror('오류', f'매칭 중 오류 발생:\n{str(e)}\n{traceback.format_exc()}')

    def show_custom_message(self, title, msg, file_path=None):
        import tkinter as tk
        from tkinter import messagebox
        import os
        win = tk.Toplevel(self.root)
        win.title(title)
        win.resizable(False, False)
        win.configure(bg='white')
        # msg가 None이면 기본 메시지로 대체
        if not msg:
            msg = '분류가 완료되었습니다!'
        # 메시지 분리 (파일 경로가 있으면 따로)
        if isinstance(msg, str) and ' | 저장 파일: ' in msg:
            main_msg, file_msg = msg.split(' | 저장 파일: ', 1)
        else:
            main_msg, file_msg = str(msg), None
        frame = tk.Frame(win, bg='white')
        frame.pack(padx=24, pady=18)
        label = tk.Label(frame, text=main_msg, font=('맑은 고딕', 12, 'bold'), bg='white', fg='#222', wraplength=400, justify='left')
        label.pack(pady=(0, 10))
        if file_msg:
            file_label = tk.Label(frame, text=f"저장 파일: {file_msg}", font=('맑은 고딕', 10), bg='white', fg='#1976d2', wraplength=400, justify='left')
            file_label.pack(pady=(0, 10))
            def copy_path():
                win.clipboard_clear()
                win.clipboard_append(file_msg)
                messagebox.showinfo('복사됨', '파일 경로가 복사되었습니다!')
            copy_btn = tk.Button(frame, text='경로 복사', command=copy_path, font=('맑은 고딕', 10))
            copy_btn.pack(pady=(0, 10))
        def on_ok():
            win.destroy()
            # 확인 버튼을 눌렀을 때만 폴더 열기
            if file_path:
                folder = os.path.dirname(file_path)
                if os.path.exists(folder):
                    try:
                        os.startfile(folder)
                    except Exception as e:
                        messagebox.showerror('폴더 열기 오류', f'폴더를 열 수 없습니다: {e}')
                else:
                    messagebox.showerror('폴더 열기 오류', f'폴더가 존재하지 않습니다: {folder}')
        ok_btn = tk.Button(frame, text='확인', command=on_ok, font=('맑은 고딕', 11), width=10, bg='#03C75A', fg='white', relief='raised')
        ok_btn.pack(pady=(5, 0))
        # 창을 프로그램 창 중앙에 배치 (winfo_rootx/y 사용)
        self.root.update_idletasks()
        win.update_idletasks()
        w = win.winfo_width()
        h = win.winfo_height()
        root_x = self.root.winfo_rootx()
        root_y = self.root.winfo_rooty()
        root_w = self.root.winfo_width()
        root_h = self.root.winfo_height()
        x = root_x + (root_w - w) // 2
        y = root_y + (root_h - h) // 2
        win.geometry(f'+{x}+{y}')

def main():
    try:
        app = DeliveryClassifierGUI()
        app.root.mainloop()
    except Exception as e:
        import traceback
        print('예외 발생:', e)
        print(traceback.format_exc())

if __name__ == "__main__":
    main() 